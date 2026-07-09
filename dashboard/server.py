#!/usr/bin/env python3
"""Trading Bot Control Dashboard — Flask API + static UI."""

from __future__ import annotations

import hashlib
import ast
import csv
import io
import json
import os
import re
import sqlite3
import secrets
import subprocess
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict, deque
from datetime import datetime, timedelta, timezone
from functools import wraps
from pathlib import Path
from types import SimpleNamespace

try:
    import psutil
except ImportError:  # pragma: no cover - fallback keeps the dashboard alive without the optional package
    psutil = None
from flask import Flask, Response, jsonify, request, send_file, send_from_directory, session
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Paths ─────────────────────────────────────────────────────────────
BOT_ROOT = Path(os.environ.get("BOT_ROOT", "/opt/trading-bot"))
DATA_ROOT = Path(os.environ.get("BOT_DATA_ROOT", str(BOT_ROOT)))
ENV_FILE = Path(os.environ.get("ENV_FILE", str(BOT_ROOT / ".env")))
STATE_FILE = BOT_ROOT / "engine_state.json"
SIGNAL_LOG = Path(os.environ.get("SIGNAL_LOG_FILE", str(DATA_ROOT / "signal_log.txt")))
SIGNAL_QUEUE = Path(os.environ.get("SIGNAL_QUEUE_FILE", str(DATA_ROOT / "signal_queue.json")))
TELEGRAM_DELIVERY_LOG = Path(
    os.environ.get("TELEGRAM_DELIVERY_LOG", str(DATA_ROOT / "telegram_delivery.log"))
)
OUTCOMES_LOG = Path(os.environ.get("SIGNAL_OUTCOMES_LOG", str(DATA_ROOT / "signal_outcomes.log")))
AUDIT_LOG = Path(os.environ.get("DASHBOARD_AUDIT_LOG", str(DATA_ROOT / "dashboard_audit.log")))
STRATEGY_DIR = BOT_ROOT / "strategy"
STRATEGY_UPLOADS = STRATEGY_DIR / "uploads"
STRATEGY_ACTIVE = STRATEGY_DIR / "active.mq5"
STRATEGY_LEGACY = BOT_ROOT / "SignalBot_MultiIndicator_MT5.mq5"
STRATEGY_MANIFEST = STRATEGY_DIR / "manifest.json"
STRATEGY_MAX_BYTES = 2 * 1024 * 1024
PM2_EVENT_LOG = Path(os.environ.get("PM2_LOG_PATH", str(Path.home() / ".pm2" / "pm2.log")))
METRICS_HISTORY: deque = deque(maxlen=720)
_LAST_ALERT_TS: dict[str, float] = {}
_LAST_SIGNAL_KEY: str | None = None
_MAINTENANCE_STATE: dict[str, bool] = {"paused_by_schedule": False}
_signals_full_cache: dict = {"mtime": 0.0, "signals": []}
_telegram_cache: dict = {"mtime": 0.0, "days": 0, "limit": 0, "entries": []}
_pm2_cache: dict = {"ts": 0.0, "data": []}
_system_micro_cache: dict = {"ts": 0.0, "data": None}
_enriched_full_cache: dict = {"sig_mtime": 0.0, "tg_mtime": 0.0, "enriched": []}
_cpu_primed = False
PM2_LOG_DIR = Path(os.environ.get("PM2_LOG_DIR", "/root/.pm2/logs"))
ENGINE_LOG = PM2_LOG_DIR / "signal-engine-error.log"
STATIC_DIR = Path(__file__).parent / "static"
VERSION_FILE = Path(__file__).parent / "version.json"
CLIENT_DIAGNOSTIC_LOG = Path(__file__).parent / "client-diagnostics.log"
CACHE_DB = Path(os.environ.get("DASHBOARD_CACHE_DB", str(BOT_ROOT / ".cache" / "dashboard-cache.sqlite3")))

PROCESSES = ("signal-engine", "signal-server")
DISPLAY_PROCESSES = ("signal-engine", "signal-server", "dashboard")
SECRET_KEYS = (
    "TWELVE_DATA_API_KEY",
    "TELEGRAM_BOT_TOKEN",
    "TELEGRAM_CHAT_ID",
    "DASHBOARD_PASSWORD",
    "DASHBOARD_TOKEN",
    "GH_PAT",
    "GITHUB_TOKEN",
)

OPS_CONFIG_KEYS = {
    "ALERT_CPU_THRESHOLD",
    "ALERT_RAM_THRESHOLD",
    "ALERT_DISK_THRESHOLD",
    "WEBHOOK_DISCORD_URL",
    "WEBHOOK_SLACK_URL",
    "WEBHOOK_ON_SIGNAL",
    "MAINTENANCE_ENABLED",
    "MAINTENANCE_WINDOW",
}

BACKUP_KEYS = {
    "SYMBOLS",
    "MIN_SCORE",
    "POLL_SECONDS",
    "FACEBOOK_ENABLE",
    "DATA_PROVIDER",
    "ENGINE_DEBUG",
    "SEND_STARTUP_MESSAGE",
    "NOTIFICATIONS_PAUSED",
    "SIGNAL_COOLDOWN_SECONDS",
    *OPS_CONFIG_KEYS,
}

def _load_dotenv() -> None:
    if not ENV_FILE.exists():
        return
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        os.environ.setdefault(key.strip(), val.strip())


_load_dotenv()


def _get_secret_key() -> str:
    """Stable secret shared across all gunicorn workers."""
    key = os.environ.get("DASHBOARD_SECRET")
    if key:
        return key
    secret_file = BOT_ROOT / ".dashboard_secret"
    if secret_file.exists():
        return secret_file.read_text().strip()
    key = secrets.token_hex(32)
    secret_file.write_text(key)
    return key


app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="/static")
app.secret_key = _get_secret_key()
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)

if os.environ.get("BEHIND_PROXY", "0") == "1":
    app.config["SESSION_COOKIE_SECURE"] = True
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)


def _dashboard_username() -> str:
    return os.environ.get("DASHBOARD_USERNAME", "admin")


def _dashboard_password() -> str:
    return os.environ.get("DASHBOARD_PASSWORD", "tradingbot2026")


_GIT_REVISION_CACHE: dict[str, str | None] = {"rev": None}


def _git_revision() -> str | None:
    cached = _GIT_REVISION_CACHE.get("rev")
    if cached is not None:
        return cached
    repo_root = Path(__file__).resolve().parents[1]
    try:
        rev = subprocess.run(
            ["git", "-C", str(repo_root), "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            check=True,
        ).stdout.strip()
        _GIT_REVISION_CACHE["rev"] = rev or None
    except Exception:
        _GIT_REVISION_CACHE["rev"] = None
    return _GIT_REVISION_CACHE["rev"]


def _dashboard_version() -> dict:
    default = {"major": 2, "minor": 17, "patch": 0, "label": "v2.17", "released": "", "history": []}
    if not VERSION_FILE.exists():
        default["revision"] = _git_revision()
        return default
    try:
        data = json.loads(VERSION_FILE.read_text())
        major = int(data.get("major", default["major"]))
        minor = int(data.get("minor", default["minor"]))
        patch = int(data.get("patch", default["patch"]))
        label = data.get("label") or f"v{major}.{minor}"
        return {
            "major": major,
            "minor": minor,
            "patch": patch,
            "label": label,
            "full": f"{major}.{minor}.{patch}",
            "released": data.get("released", ""),
            "history": data.get("history", []),
            "revision": _git_revision(),
        }
    except (json.JSONDecodeError, TypeError, ValueError):
        default["revision"] = _git_revision()
        return default


def auth_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("authenticated"):
            return f(*args, **kwargs)
        token = request.headers.get("X-Dashboard-Token")
        if token and token == os.environ.get("DASHBOARD_TOKEN"):
            return f(*args, **kwargs)
        return jsonify({"error": "Unauthorized"}), 401

    return wrapper


def _run(cmd: list[str], timeout: int = 15) -> tuple[int, str, str]:
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return r.returncode, r.stdout.strip(), r.stderr.strip()
    except subprocess.TimeoutExpired:
        return 1, "", "timeout"
    except FileNotFoundError:
        return 1, "", "command not found"


def _pm2_list() -> list[dict]:
    code, out, _ = _run(["pm2", "jlist"])
    if code != 0 or not out:
        return []
    try:
        return json.loads(out)
    except json.JSONDecodeError:
        return []


def _pm2_list_cached(max_age: float = 2.0) -> list[dict]:
    now = time.time()
    if _pm2_cache["data"] and now - _pm2_cache["ts"] < max_age:
        return _pm2_cache["data"]
    data = _pm2_list()
    _pm2_cache["ts"] = now
    _pm2_cache["data"] = data
    return data


def _pm2_by_name() -> dict[str, dict]:
    return {p.get("name"): p for p in _pm2_list_cached() if p.get("name")}


def _proc_info(name: str, pm2_map: dict[str, dict] | None = None) -> dict:
    p = (pm2_map or _pm2_by_name()).get(name)
    if p:
        env = p.get("pm2_env") or {}
        monit = p.get("monit") or {}
        return {
            "name": name,
            "status": env.get("status", "unknown"),
            "pid": p.get("pid", 0),
            "uptime": env.get("pm_uptime"),
            "restarts": env.get("restart_time", 0),
            "memory_mb": round((monit.get("memory") or 0) / 1024 / 1024, 1),
            "cpu": monit.get("cpu", 0),
        }
    return {"name": name, "status": "not_found", "pid": 0, "restarts": 0, "memory_mb": 0, "cpu": 0}


def _uptime_str(ms: int | None) -> str:
    if not ms:
        return "—"
    secs = max(0, int(time.time() * 1000 - ms) // 1000)
    d, rem = divmod(secs, 86400)
    h, rem = divmod(rem, 3600)
    m, s = divmod(rem, 60)
    parts = []
    if d:
        parts.append(f"{d}d")
    if h:
        parts.append(f"{h}h")
    if m:
        parts.append(f"{m}m")
    if not parts:
        parts.append(f"{s}s")
    return " ".join(parts)


def _read_json(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text())
    except Exception:
        return None


def _cache_db() -> sqlite3.Connection:
    CACHE_DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(CACHE_DB, timeout=0.4)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=400")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cache (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            expires_at REAL NOT NULL,
            updated_at REAL NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cache_expires ON cache(expires_at)")
    return conn


def _cache_get(key: str):
    now = time.time()
    try:
        with _cache_db() as conn:
            row = conn.execute(
                "SELECT value, expires_at FROM cache WHERE key = ?",
                (key,),
            ).fetchone()
            if not row:
                return None
            if row["expires_at"] <= now:
                conn.execute("DELETE FROM cache WHERE key = ?", (key,))
                return None
            return json.loads(row["value"])
    except Exception:
        return None


def _cache_set(key: str, value, ttl_seconds: int) -> None:
    expires_at = time.time() + max(ttl_seconds, 1)
    payload = json.dumps(value, ensure_ascii=False, default=str)
    try:
        with _cache_db() as conn:
            conn.execute(
                """
                INSERT INTO cache (key, value, expires_at, updated_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET
                    value = excluded.value,
                    expires_at = excluded.expires_at,
                    updated_at = excluded.updated_at
                """,
                (key, payload, expires_at, time.time()),
            )
    except Exception:
        pass


def _cache_delete_prefix(prefix: str) -> None:
    try:
        with _cache_db() as conn:
            conn.execute("DELETE FROM cache WHERE key LIKE ?", (f"{prefix}%",))
    except Exception:
        pass


def _cache_json(key: str, ttl_seconds: int, loader):
    cached = _cache_get(key)
    if cached is not None:
        return cached
    value = loader()
    _cache_set(key, value, ttl_seconds)
    return value


def _invalidate_dashboard_cache() -> None:
    for prefix in (
        "bootstrap:",
        "signals:v1:",
        "reports:summary:v1:",
        "telegram:log:v1:",
        "telegram:summary:v1:",
        "analytics:symbols:v1:",
        "analytics:hourly:v1:",
    ):
        _cache_delete_prefix(prefix)


def _parse_env() -> dict[str, str]:
    if not ENV_FILE.exists():
        return {}
    out: dict[str, str] = {}
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        out[k.strip()] = v.strip()
    return out


def _write_env(updates: dict[str, str]) -> None:
    lines: list[str] = []
    seen: set[str] = set()
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            stripped = line.strip()
            if stripped and not stripped.startswith("#") and "=" in stripped:
                k = stripped.split("=", 1)[0].strip()
                if k in updates:
                    lines.append(f"{k}={updates[k]}")
                    seen.add(k)
                    continue
            lines.append(line)
    for k, v in updates.items():
        if k not in seen:
            lines.append(f"{k}={v}")
    ENV_FILE.write_text("\n".join(lines) + "\n")


def _notifications_paused() -> bool:
    return _parse_env().get("NOTIFICATIONS_PAUSED", "0") == "1"


def _stop_signal_engine() -> tuple[int, str, str]:
    return _run(["pm2", "stop", "signal-engine"])


def _start_signal_engine(*, force: bool = False) -> tuple[bool, str]:
    """Start engine unless notifications are paused (unless force=True)."""
    if _notifications_paused() and not force:
        _stop_signal_engine()
        return False, "نوتیفیکیشن متوقف است — ابتدا «ادامه نوتیفیکیشن» را بزنید"
    code, out, err = _run(["pm2", "start", "signal-engine"])
    if code != 0:
        return False, err or out or "pm2 start failed"
    return True, out or "started"


def _restart_signal_engine(*, force: bool = False) -> tuple[bool, str, bool]:
    """Restart engine. Returns (ok, message, skipped_due_to_pause)."""
    if _notifications_paused() and not force:
        _stop_signal_engine()
        _run(["pm2", "save"])
        return True, "استراتژی ذخیره شد — موتور متوقف ماند (نوتیفیکیشن pause)", True
    code, out, err = _run(["pm2", "restart", "signal-engine"])
    _run(["pm2", "save"])
    if code != 0:
        return False, err or out or "pm2 restart failed", False
    return True, out or "restarted", False


def _mask(val: str) -> str:
    if not val:
        return ""
    if len(val) <= 8:
        return "••••••••"
    return val[:4] + "•" * (len(val) - 8) + val[-4:]


def _enriched_source_mtime() -> tuple[float, float]:
    sig = SIGNAL_LOG.stat().st_mtime if SIGNAL_LOG.exists() else 0.0
    tg = TELEGRAM_DELIVERY_LOG.stat().st_mtime if TELEGRAM_DELIVERY_LOG.exists() else 0.0
    return sig, tg


def _get_enriched_all() -> list[dict]:
    """Enrich full signal log once; reuse until signal or telegram log changes."""
    sig_m, tg_m = _enriched_source_mtime()
    cached = _enriched_full_cache
    if cached["sig_mtime"] == sig_m and cached["tg_mtime"] == tg_m and cached["enriched"]:
        return cached["enriched"]
    all_signals = _get_all_signals_full()
    telegram = _parse_telegram_deliveries(days=30, limit=5000)
    enriched = _enrich_signals(all_signals, telegram)
    cached["sig_mtime"] = sig_m
    cached["tg_mtime"] = tg_m
    cached["enriched"] = enriched
    return enriched


def _get_enriched_signals(days: int | None = 30) -> list[dict]:
    enriched = _get_enriched_all()
    if days is None:
        return list(enriched)
    cutoff = datetime.now() - timedelta(days=days)
    return [s for s in enriched if _signal_after(s, cutoff)]


def _parse_all_signals(days: int | None = None) -> list[dict]:
    all_sigs = _get_all_signals_full()
    if days is None:
        return list(all_sigs)
    cutoff = datetime.now() - timedelta(days=days)
    filtered: list[dict] = []
    for s in all_sigs:
        ts_str = s.get("timestamp", "")
        if not ts_str:
            continue
        try:
            if datetime.strptime(ts_str[:19], "%Y-%m-%d %H:%M:%S") >= cutoff:
                filtered.append(s)
        except ValueError:
            filtered.append(s)
    return filtered


def _get_all_signals_full() -> list[dict]:
    """Parse signal log once; reuse until file mtime changes."""
    if not SIGNAL_LOG.exists():
        return []
    mtime = SIGNAL_LOG.stat().st_mtime
    if _signals_full_cache["mtime"] == mtime and _signals_full_cache["signals"]:
        return _signals_full_cache["signals"]
    signals: list[dict] = []
    pattern = re.compile(
        r"\[(?P<ts>[^\]]+)\] Signal saved → (?P<data>.+)$"
    )
    for line in SIGNAL_LOG.read_text().splitlines():
        m = pattern.search(line)
        if not m:
            continue
        ts_str = m.group("ts")
        try:
            data = ast.literal_eval(m.group("data"))
        except Exception:
            continue
        signals.append({"timestamp": ts_str, **data})
    signals = list(reversed(signals))
    _signals_full_cache["mtime"] = mtime
    _signals_full_cache["signals"] = signals
    return signals


def _parse_signals(limit: int = 50, days: int | None = None) -> list[dict]:
    pool = _parse_all_signals(days=days)
    return pool[:limit]


_OUTCOME_ENGINE_RES = (
    re.compile(
        r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ .*?(?P<outcome>TP1|TP2|TP|SL|STOP.?LOSS|TAKE.?PROFIT).*?hit.*?(?P<symbol>[\w/]+)",
        re.I,
    ),
    re.compile(
        r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ .*?(?P<symbol>[\w/]+).*?(?P<outcome>TP1|TP2|TP|SL).*?(?:hit|reached|triggered)",
        re.I,
    ),
)


def _parse_outcomes(days: int | None = 30) -> list[dict]:
    """Load trade outcomes from JSONL log and engine log patterns."""
    outcomes: list[dict] = []
    cutoff = datetime.now() - timedelta(days=days) if days else None

    if OUTCOMES_LOG.exists():
        for line in reversed(OUTCOMES_LOG.read_text(encoding="utf-8", errors="replace").splitlines()):
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            ts = str(row.get("ts") or row.get("timestamp") or "")
            if cutoff and _parse_ts(ts) and _parse_ts(ts) < cutoff:
                continue
            raw_out = str(row.get("outcome") or row.get("result") or "open").lower()
            if raw_out in ("tp", "take_profit", "takeprofit"):
                raw_out = "tp1"
            elif raw_out.startswith("tp2"):
                raw_out = "tp2"
            elif raw_out.startswith("tp1"):
                raw_out = "tp1"
            elif raw_out in ("sl", "stop_loss", "stoploss", "loss"):
                raw_out = "sl"
            elif raw_out in ("expired", "timeout"):
                raw_out = "expired"
            else:
                raw_out = "open"
            outcomes.append({
                "timestamp": ts[:19],
                "symbol": _normalize_symbol(str(row.get("symbol", ""))),
                "direction": str(row.get("direction", "")).upper(),
                "entry": row.get("entry"),
                "outcome": raw_out,
                "exit_price": row.get("exit_price"),
                "source": "outcomes_log",
            })

    if ENGINE_LOG.exists():
        seen_engine: set[str] = set()
        for line in reversed(ENGINE_LOG.read_text(encoding="utf-8", errors="replace").splitlines()):
            for pat in _OUTCOME_ENGINE_RES:
                m = pat.search(line)
                if not m:
                    continue
                ts_str = m.group("ts")
                if cutoff and _parse_ts(ts_str) and _parse_ts(ts_str) < cutoff:
                    continue
                oc_raw = m.group("outcome").upper()
                if "TP2" in oc_raw:
                    oc = "tp2"
                elif "TP" in oc_raw:
                    oc = "tp1"
                else:
                    oc = "sl"
                key = f"{ts_str}|{m.group('symbol')}|{oc}"
                if key in seen_engine:
                    continue
                seen_engine.add(key)
                outcomes.append({
                    "timestamp": ts_str,
                    "symbol": _normalize_symbol(m.group("symbol")),
                    "direction": "",
                    "entry": None,
                    "outcome": oc,
                    "exit_price": None,
                    "source": "engine_log",
                })
                break

    return outcomes


def _attach_outcomes(signals: list[dict], outcomes: list[dict]) -> list[dict]:
    for sig in signals:
        sym = _normalize_symbol(str(sig.get("symbol", "")))
        direction = str(sig.get("direction", "")).upper()
        ts = _parse_ts(sig.get("timestamp", ""))
        match = None
        best_delta = 999999.0
        for o in outcomes:
            if _normalize_symbol(str(o.get("symbol", ""))) != sym:
                continue
            o_dir = str(o.get("direction", "")).upper()
            if o_dir and direction and o_dir != direction:
                continue
            o_ts = _parse_ts(o.get("timestamp", ""))
            if ts and o_ts:
                delta = abs((ts - o_ts).total_seconds())
                if delta > 86400:
                    continue
                if delta < best_delta:
                    best_delta = delta
                    match = o
            elif not ts:
                match = o
                break
        if match and match.get("outcome") not in (None, "open"):
            sig["outcome"] = match["outcome"]
            sig["outcome_source"] = match.get("source")
            sig["exit_price"] = match.get("exit_price")
        else:
            sig["outcome"] = "open"
            sig["outcome_source"] = None
            sig["exit_price"] = None
    return signals


def _outcome_summary(signals: list[dict]) -> dict:
    wins = [s for s in signals if s.get("outcome") in ("tp1", "tp2")]
    losses = [s for s in signals if s.get("outcome") == "sl"]
    open_ = [s for s in signals if s.get("outcome") in (None, "open")]
    expired = [s for s in signals if s.get("outcome") == "expired"]
    closed = len(wins) + len(losses)
    today = datetime.now().strftime("%Y-%m-%d")
    today_sigs = [s for s in signals if str(s.get("timestamp", "")).startswith(today)]
    today_wins = [s for s in today_sigs if s.get("outcome") in ("tp1", "tp2")]
    today_losses = [s for s in today_sigs if s.get("outcome") == "sl"]
    today_closed = len(today_wins) + len(today_losses)
    return {
        "wins": len(wins),
        "losses": len(losses),
        "open": len(open_),
        "expired": len(expired),
        "closed": closed,
        "win_rate": round(len(wins) / closed * 100, 1) if closed else None,
        "today_wins": len(today_wins),
        "today_losses": len(today_losses),
        "today_closed": today_closed,
        "today_win_rate": round(len(today_wins) / today_closed * 100, 1) if today_closed else None,
    }


def _log_telegram_delivery(**kwargs) -> None:
    try:
        sys.path.insert(0, str(BOT_ROOT))
        from telegram_logger import log_telegram_delivery

        log_telegram_delivery(**kwargs)
    except Exception:
        TELEGRAM_DELIVERY_LOG.parent.mkdir(parents=True, exist_ok=True)
        record = {
            "ts": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            **kwargs,
        }
        with TELEGRAM_DELIVERY_LOG.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")


def _telegram_send(text: str) -> dict:
    env = _parse_env()
    token = env.get("TELEGRAM_BOT_TOKEN", "").strip()
    chat_id = env.get("TELEGRAM_CHAT_ID", "").strip()
    if not token or not chat_id:
        return {"ok": False, "error": "تلگرام تنظیم نشده (TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID)", "http_status": None}
    payload = json.dumps({"chat_id": chat_id, "text": text, "parse_mode": "HTML"}).encode("utf-8")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            body = json.loads(resp.read().decode("utf-8"))
            ok = bool(body.get("ok"))
            err = None if ok else str(body.get("description") or body)
            return {"ok": ok, "error": err, "http_status": resp.status, "response": body}
    except urllib.error.HTTPError as exc:
        try:
            body = json.loads(exc.read().decode("utf-8"))
            err = str(body.get("description") or body)
        except Exception:
            err = str(exc)
        return {"ok": False, "error": err, "http_status": exc.code}
    except Exception as exc:
        return {"ok": False, "error": str(exc), "http_status": None}


def _format_signal_telegram_message(sig: dict) -> str:
    sym = sig.get("symbol", "?")
    direction = str(sig.get("direction", "")).upper()
    emoji = "🟢" if direction == "BUY" else "🔴"
    lines = [
        f"{emoji} <b>{direction} {sym}</b>",
        f"Entry: <code>{sig.get('entry', '—')}</code>",
        f"SL: <code>{sig.get('sl', '—')}</code>",
        f"TP1: <code>{sig.get('tp1', '—')}</code>",
    ]
    if sig.get("tp2"):
        lines.append(f"TP2: <code>{sig.get('tp2')}</code>")
    if sig.get("rr"):
        lines.append(f"RR: {sig.get('rr')}")
    if sig.get("score") is not None:
        lines.append(f"Score: {sig.get('score')}")
    if sig.get("basis"):
        lines.append(f"\n{sig.get('basis')}")
    return "\n".join(lines)


def _find_signal_for_retry(symbol: str, timestamp: str = "", direction: str = "") -> dict | None:
    sym = _normalize_symbol(symbol)
    direction = direction.upper()
    candidates = _parse_all_signals(days=30)
    best = None
    best_delta = 999999.0
    target_ts = _parse_ts(timestamp) if timestamp else None
    for sig in candidates:
        if _normalize_symbol(str(sig.get("symbol", ""))) != sym:
            continue
        if direction and str(sig.get("direction", "")).upper() != direction:
            continue
        if target_ts:
            sig_ts = _parse_ts(sig.get("timestamp", ""))
            if not sig_ts:
                continue
            delta = abs((target_ts - sig_ts).total_seconds())
            if delta < best_delta:
                best_delta = delta
                best = sig
        else:
            return sig
    return best if best_delta <= 600 else None


def _enrich_signals(signals: list[dict], telegram_entries: list[dict]) -> list[dict]:
    """Attach telegram delivery status and quality flags to each saved signal."""
    tg_signals = [e for e in telegram_entries if e.get("message_type", "signal") == "signal"]
    enriched: list[dict] = []
    seen_recent: list[tuple[str, str, datetime]] = []

    for sig in signals:
        row = dict(sig)
        sym = _normalize_symbol(str(sig.get("symbol", "")))
        direction = str(sig.get("direction", "")).upper()
        ts = _parse_ts(sig.get("timestamp", ""))

        match = None
        best_delta = 999999.0
        for t in tg_signals:
            if _normalize_symbol(str(t.get("symbol", ""))) != sym:
                continue
            t_dir = str(t.get("direction", "")).upper()
            if t_dir and direction and t_dir != direction:
                continue
            t_ts = _parse_ts(t.get("timestamp", ""))
            if ts and t_ts:
                delta = abs((ts - t_ts).total_seconds())
                if delta > 300:
                    continue
                if t.get("entry") and sig.get("entry"):
                    try:
                        e1 = float(t["entry"])
                        e2 = float(sig["entry"])
                        if abs(e1 - e2) / max(abs(e2), 1e-9) > 0.002:
                            continue
                    except (TypeError, ValueError):
                        pass
                if delta < best_delta:
                    best_delta = delta
                    match = t
            elif not ts:
                match = t
                break

        if match:
            row["delivery_status"] = "sent" if match.get("ok") else "failed"
            row["telegram_ok"] = bool(match.get("ok"))
            row["telegram_detail"] = match.get("detail") or match.get("error") or ""
            row["score"] = match.get("score") or sig.get("score")
            row["telegram_source"] = match.get("source")
        else:
            row["delivery_status"] = "unsent"
            row["telegram_ok"] = None
            row["telegram_detail"] = "ارسال تلگرام تأیید نشده"
            row["score"] = sig.get("score")

        duplicate = False
        if ts:
            for psym, pdir, pts in seen_recent:
                if psym == sym and pdir == direction and abs((ts - pts).total_seconds()) < 1800:
                    duplicate = True
                    break
            seen_recent.append((sym, direction, ts))
        row["duplicate"] = duplicate
        if duplicate and row["delivery_status"] == "sent":
            row["quality"] = "duplicate"
        elif row["delivery_status"] == "failed":
            row["quality"] = "mistaken"
        elif row["delivery_status"] == "unsent":
            row["quality"] = "pending"
        else:
            row["quality"] = "ok"

        enriched.append(row)

    enriched.sort(key=lambda s: s.get("timestamp", ""), reverse=True)
    outcomes = _parse_outcomes(days=90)
    return _attach_outcomes(enriched, outcomes)


def _signals_page_summary(enriched: list[dict]) -> dict:
    today = datetime.now().strftime("%Y-%m-%d")
    sent = [s for s in enriched if s.get("delivery_status") == "sent"]
    failed = [s for s in enriched if s.get("delivery_status") == "failed"]
    unsent = [s for s in enriched if s.get("delivery_status") == "unsent"]
    dupes = [s for s in enriched if s.get("duplicate")]
    today_list = [s for s in enriched if str(s.get("timestamp", "")).startswith(today)]
    by_symbol: dict[str, int] = {}
    for s in enriched:
        sym = _normalize_symbol(str(s.get("symbol", "?")))
        by_symbol[sym] = by_symbol.get(sym, 0) + 1
    daily: dict[str, dict] = {}
    for s in enriched:
        day = str(s.get("timestamp", ""))[:10]
        if not day:
            continue
        if day not in daily:
            daily[day] = {"date": day, "total": 0, "sent": 0, "failed": 0, "unsent": 0}
        daily[day]["total"] += 1
        st = s.get("delivery_status", "unsent")
        if st in daily[day]:
            daily[day][st] += 1
    daily_list = sorted(daily.values(), key=lambda d: d["date"])[-30:]
    total = len(enriched) or 1
    oc = _outcome_summary(enriched)
    return {
        "total": len(enriched),
        "sent": len(sent),
        "failed": len(failed),
        "unsent": len(unsent),
        "duplicates": len(dupes),
        "mistaken": len(failed) + len(dupes),
        "today": len(today_list),
        "today_sent": len([s for s in today_list if s.get("delivery_status") == "sent"]),
        "delivery_rate": round(len(sent) / total * 100, 1),
        "by_symbol": by_symbol,
        "by_direction": {
            "BUY": len([s for s in enriched if str(s.get("direction", "")).upper() == "BUY"]),
            "SELL": len([s for s in enriched if str(s.get("direction", "")).upper() == "SELL"]),
        },
        "daily": daily_list,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "outcomes": oc,
    }


def _normalize_symbol(sym: str) -> str:
    sym = (sym or "").strip().upper()
    if not sym:
        return "?"
    if "/" in sym:
        return sym
    if len(sym) == 6:
        return f"{sym[:3]}/{sym[3:]}"
    if sym.endswith("USD") and len(sym) > 3:
        return f"{sym[:-3]}/USD"
    return sym


def _parse_ts(ts_str: str) -> datetime | None:
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S UTC"):
        try:
            return datetime.strptime(ts_str[:19], fmt[:19])
        except ValueError:
            continue
    return None


def _within_days(ts_str: str, days: int | None) -> bool:
    if days is None:
        return True
    ts = _parse_ts(ts_str)
    if not ts:
        return True
    return ts >= datetime.now() - timedelta(days=days)


_TELEGRAM_SENT_RE = re.compile(
    r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ \w+ Signal sent "
    r"(?P<symbol>[\w/]+) (?P<direction>BUY|SELL)(?: score=(?P<score>\d+))?(?: entry=(?P<entry>[\d.]+))?"
)
_TELEGRAM_FAIL_RES = (
    re.compile(
        r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ ERROR (?P<msg>Telegram(?: send)? failed[^\\n]*?)"
        r"(?:\s+symbol[=:]?(?P<symbol>[\w/]+))?",
        re.I,
    ),
    re.compile(
        r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ ERROR (?P<msg>.*Telegram.*)",
        re.I,
    ),
    re.compile(
        r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ WARNING (?P<msg>Telegram[^\\n]+)",
        re.I,
    ),
)
_TELEGRAM_STARTUP_RE = re.compile(
    r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),\d+ \w+ (?P<msg>.*(?:startup|Startup).*(?:Telegram|telegram|sent).*)",
    re.I,
)


def _parse_telegram_jsonl(days: int | None = None, limit: int = 500) -> list[dict]:
    if not TELEGRAM_DELIVERY_LOG.exists():
        return []
    entries: list[dict] = []
    for line in reversed(TELEGRAM_DELIVERY_LOG.read_text(encoding="utf-8", errors="replace").splitlines()):
        line = line.strip()
        if not line:
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        ts = row.get("ts", "")
        if not _within_days(ts, days):
            continue
        ok = bool(row.get("ok"))
        entries.append(
            {
                "timestamp": ts.replace(" UTC", ""),
                "symbol": _normalize_symbol(str(row.get("symbol", "?"))),
                "direction": str(row.get("direction", "")).upper(),
                "status": "ok" if ok else "failed",
                "ok": ok,
                "score": row.get("score"),
                "entry": row.get("entry"),
                "error": row.get("error"),
                "http_status": row.get("http_status"),
                "message_type": row.get("message_type", "signal"),
                "source": "delivery_log",
                "detail": row.get("error") or ("ارسال موفق به تلگرام" if ok else "ارسال ناموفق"),
            }
        )
        if len(entries) >= limit:
            break
    return entries


def _parse_engine_telegram_logs(days: int | None = None, limit: int = 1000) -> list[dict]:
    if not ENGINE_LOG.exists():
        return []
    entries: list[dict] = []
    for line in reversed(ENGINE_LOG.read_text(encoding="utf-8", errors="replace").splitlines()):
        m = _TELEGRAM_SENT_RE.search(line)
        if m:
            ts = m.group("ts")
            if not _within_days(ts, days):
                continue
            entries.append(
                {
                    "timestamp": ts,
                    "symbol": _normalize_symbol(m.group("symbol")),
                    "direction": m.group("direction"),
                    "status": "ok",
                    "ok": True,
                    "score": int(m.group("score")) if m.group("score") else None,
                    "entry": m.group("entry"),
                    "error": None,
                    "http_status": 200,
                    "message_type": "signal",
                    "source": "engine_log",
                    "detail": "سیگنال با موفقیت به تلگرام ارسال شد",
                }
            )
            if len(entries) >= limit:
                break
            continue

        for fail_re in _TELEGRAM_FAIL_RES:
            fm = fail_re.search(line)
            if fm:
                ts = fm.group("ts")
                if not _within_days(ts, days):
                    continue
                sym = fm.groupdict().get("symbol")
                entries.append(
                    {
                        "timestamp": ts,
                        "symbol": _normalize_symbol(sym) if sym else "—",
                        "direction": "",
                        "status": "failed",
                        "ok": False,
                        "score": None,
                        "entry": None,
                        "error": fm.group("msg").strip(),
                        "http_status": None,
                        "message_type": "signal",
                        "source": "engine_log",
                        "detail": fm.group("msg").strip(),
                    }
                )
                break

        sm = _TELEGRAM_STARTUP_RE.search(line)
        if sm:
            ts = sm.group("ts")
            if not _within_days(ts, days):
                continue
            ok = "fail" not in sm.group("msg").lower()
            entries.append(
                {
                    "timestamp": ts,
                    "symbol": "—",
                    "direction": "",
                    "status": "ok" if ok else "failed",
                    "ok": ok,
                    "score": None,
                    "entry": None,
                    "error": None if ok else sm.group("msg"),
                    "message_type": "startup",
                    "source": "engine_log",
                    "detail": sm.group("msg").strip(),
                }
            )
        if len(entries) >= limit:
            break
    return entries


def _merge_telegram_entries(*sources: list[dict]) -> list[dict]:
    seen: set[str] = set()
    merged: list[dict] = []
    for source in sources:
        for row in source:
            key = f"{row.get('timestamp')}|{row.get('symbol')}|{row.get('direction')}|{row.get('status')}|{row.get('message_type')}"
            if key in seen:
                continue
            seen.add(key)
            merged.append(row)
    merged.sort(key=lambda r: r.get("timestamp", ""), reverse=True)
    return merged


def _parse_telegram_deliveries(days: int | None = 30, limit: int = 200) -> list[dict]:
    mtime = TELEGRAM_DELIVERY_LOG.stat().st_mtime if TELEGRAM_DELIVERY_LOG.exists() else 0.0
    if (
        _telegram_cache["entries"]
        and _telegram_cache["mtime"] == mtime
        and _telegram_cache["days"] == (days or 0)
        and _telegram_cache["limit"] >= limit
    ):
        return _telegram_cache["entries"][:limit]
    jsonl = _parse_telegram_jsonl(days=days, limit=limit)
    engine = _parse_engine_telegram_logs(days=days, limit=limit * 2)
    merged = _merge_telegram_entries(jsonl, engine)[: max(limit, 500)]
    _telegram_cache.update({"mtime": mtime, "days": days or 0, "limit": max(limit, 500), "entries": merged})
    return merged[:limit]


def _telegram_summary(entries: list[dict], days: int = 30) -> dict:
    today = datetime.now().strftime("%Y-%m-%d")
    ok = [e for e in entries if e.get("ok")]
    failed = [e for e in entries if not e.get("ok")]
    signals = [e for e in entries if e.get("message_type", "signal") == "signal"]
    today_ok = sum(1 for e in ok if str(e.get("timestamp", "")).startswith(today))
    today_failed = sum(1 for e in failed if str(e.get("timestamp", "")).startswith(today))
    total_attempts = len(signals) or len(entries)
    success_rate = round(len([e for e in signals if e.get("ok")]) / max(total_attempts, 1) * 100, 1)

    daily: dict[str, dict] = {}
    for i in range(min(days, 90)):
        d = (datetime.now() - timedelta(days=min(days, 90) - 1 - i)).strftime("%Y-%m-%d")
        daily[d] = {"date": d, "ok": 0, "failed": 0, "total": 0}

    for e in entries:
        day = str(e.get("timestamp", ""))[:10]
        if day not in daily:
            continue
        daily[day]["total"] += 1
        if e.get("ok"):
            daily[day]["ok"] += 1
        else:
            daily[day]["failed"] += 1

    by_symbol: dict[str, int] = defaultdict(int)
    for e in ok:
        if e.get("symbol") and e.get("symbol") != "—":
            by_symbol[e.get("symbol", "?")] += 1

    last = entries[0] if entries else None
    return {
        "total": len(entries),
        "signals_sent": len([e for e in signals if e.get("ok")]),
        "failed": len([e for e in signals if not e.get("ok")]),
        "success_rate": success_rate,
        "today_ok": today_ok,
        "today_failed": today_failed,
        "last_delivery": last,
        "by_symbol": dict(by_symbol),
        "daily": list(daily.values()),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _daily_breakdown(signals: list[dict], days: int = 7) -> list[dict]:
    buckets: dict[str, dict] = defaultdict(lambda: {"date": "", "total": 0, "buy": 0, "sell": 0})
    for i in range(days):
        d = (datetime.now() - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
        buckets[d] = {"date": d, "total": 0, "buy": 0, "sell": 0}
    for s in signals:
        day = s.get("timestamp", "")[:10]
        if day not in buckets:
            continue
        buckets[day]["total"] += 1
        direction = s.get("direction", "").upper()
        if direction == "BUY":
            buckets[day]["buy"] += 1
        elif direction == "SELL":
            buckets[day]["sell"] += 1
    return list(buckets.values())


def _report_summary(signals: list[dict], days: int = 30) -> dict:
    stats = _signal_stats(signals)
    chart_days = min(max(days, 7), 90)
    daily = _daily_breakdown(signals, chart_days)
    active_days = max(len([d for d in daily if d["total"] > 0]), 1)
    avg_per_day = round(stats["total"] / max(days, 1), 1)
    top_symbol = max(stats["by_symbol"], key=stats["by_symbol"].get) if stats["by_symbol"] else "—"
    buy = stats["by_direction"].get("BUY", 0)
    sell = stats["by_direction"].get("SELL", 0)
    ratio = f"{round(buy / sell, 2)}:1" if sell else "—"
    procs = [_proc_info(n) for n in PROCESSES]
    total_restarts = sum(p.get("restarts", 0) for p in procs)

    telegram_entries = _parse_telegram_deliveries(days=days, limit=500)
    telegram = _telegram_summary(telegram_entries, days=days)
    telegram["configured"] = bool(_parse_env().get("TELEGRAM_BOT_TOKEN"))

    return {
        **stats,
        "days": days,
        "daily": daily,
        "avg_per_day": avg_per_day,
        "top_symbol": top_symbol,
        "buy_sell_ratio": ratio,
        "total_restarts": total_restarts,
        "telegram": telegram,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _build_xlsx(signals: list[dict], summary: dict) -> io.BytesIO:
    buf = io.BytesIO()
    if not HAS_OPENPYXL:
        return buf
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a2332")
    accent_fill = PatternFill("solid", fgColor="e8f5f0")

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Trading Bot Performance Report"])
    ws.append(["Generated", summary["generated_at"]])
    ws.append([])
    rows = [
        ("Total Signals", summary["total"]),
        ("Today", summary["today"]),
        ("Average / Day (7d)", summary["avg_per_day"]),
        ("Top Symbol", summary["top_symbol"]),
        ("BUY Signals", summary["by_direction"].get("BUY", 0)),
        ("SELL Signals", summary["by_direction"].get("SELL", 0)),
        ("Buy/Sell Ratio", summary["buy_sell_ratio"]),
        ("PM2 Restarts", summary["total_restarts"]),
        ("Telegram Sent", summary.get("telegram", {}).get("signals_sent", 0)),
        ("Telegram Failed", summary.get("telegram", {}).get("failed", 0)),
        ("Telegram Success %", summary.get("telegram", {}).get("success_rate", 0)),
    ]
    for label, val in rows:
        ws.append([label, val])
    ws.append([])
    ws.append(["Daily Breakdown (7 days)"])
    ws.append(["Date", "Total", "BUY", "SELL"])
    for row in ws[5]:
        row.font = header_font
        row.fill = header_fill
    for day in summary["daily"]:
        ws.append([day["date"], day["total"], day["buy"], day["sell"]])

    # Signals sheet
    ws2 = wb.create_sheet("Signals")
    headers = ["Timestamp", "Symbol", "Direction", "Entry", "SL", "TP1", "TP2", "RR", "Basis", "Telegram"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for s in signals:
        ws2.append([
            s.get("timestamp", ""),
            s.get("symbol", ""),
            s.get("direction", ""),
            s.get("entry", ""),
            s.get("sl", ""),
            s.get("tp1", ""),
            s.get("tp2", ""),
            s.get("rr", ""),
            s.get("basis", ""),
            s.get("telegram_status", "—"),
        ])

    # By symbol sheet
    ws3 = wb.create_sheet("By Symbol")
    ws3.append(["Symbol", "Count", "Share %"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    total = summary["total"] or 1
    for sym, count in sorted(summary["by_symbol"].items(), key=lambda x: -x[1]):
        ws3.append([sym, count, f"{round(count / total * 100, 1)}%"])

    wb.save(buf)
    buf.seek(0)
    return buf


def _signal_stats(signals: list[dict]) -> dict:
    today = datetime.now().strftime("%Y-%m-%d")
    today_count = sum(1 for s in signals if s.get("timestamp", "").startswith(today))
    by_symbol: dict[str, int] = {}
    by_dir: dict[str, int] = {"BUY": 0, "SELL": 0}
    for s in signals:
        sym = s.get("symbol", "?")
        by_symbol[sym] = by_symbol.get(sym, 0) + 1
        d = s.get("direction", "").upper()
        if d in by_dir:
            by_dir[d] += 1
    return {
        "total": len(signals),
        "today": today_count,
        "by_symbol": by_symbol,
        "by_direction": by_dir,
    }


_signal_stats_cache: dict = {"mtime": 0.0, "stats": None}


def _live_signal_stats() -> dict:
    """Full signal-log stats — cached until the log file changes."""
    if not SIGNAL_LOG.exists():
        return _signal_stats([])
    mtime = SIGNAL_LOG.stat().st_mtime
    if _signal_stats_cache["stats"] is not None and _signal_stats_cache["mtime"] == mtime:
        return _signal_stats_cache["stats"]
    stats = _signal_stats(_parse_all_signals())
    _signal_stats_cache["mtime"] = mtime
    _signal_stats_cache["stats"] = stats
    return stats


_net_prev: dict[str, float] = {"ts": 0, "sent": 0, "recv": 0}


def _fallback_cpu_percent() -> float:
    try:
        load_1m = os.getloadavg()[0]
        cores = os.cpu_count() or 1
        return round(min((load_1m / cores) * 100, 100.0), 1)
    except Exception:
        return 0.0


def _fallback_virtual_memory() -> SimpleNamespace:
    try:
        meminfo: dict[str, int] = {}
        for line in Path("/proc/meminfo").read_text().splitlines():
            key, value, *_ = line.split()
            meminfo[key.rstrip(":")] = int(value) * 1024
        total = meminfo.get("MemTotal", 0)
        available = meminfo.get("MemAvailable", meminfo.get("MemFree", 0))
        used = max(total - available, 0)
        percent = round((used / total) * 100, 1) if total else 0.0
        return SimpleNamespace(total=total, available=available, used=used, percent=percent)
    except Exception:
        return SimpleNamespace(total=0, available=0, used=0, percent=0.0)


def _fallback_disk_usage(path: str = "/") -> SimpleNamespace:
    try:
        st = os.statvfs(path)
        total = st.f_frsize * st.f_blocks
        free = st.f_frsize * st.f_bavail
        used = max(total - free, 0)
        percent = round((used / total) * 100, 1) if total else 0.0
        return SimpleNamespace(total=total, free=free, used=used, percent=percent)
    except Exception:
        return SimpleNamespace(total=0, free=0, used=0, percent=0.0)


def _fallback_net_io() -> SimpleNamespace:
    try:
        sent = 0
        recv = 0
        for line in Path("/proc/net/dev").read_text().splitlines()[2:]:
            if ":" not in line:
                continue
            iface, rest = line.split(":", 1)
            iface = iface.strip()
            if iface == "lo":
                continue
            parts = rest.split()
            if len(parts) >= 16:
                recv += int(parts[0])
                sent += int(parts[8])
        return SimpleNamespace(bytes_sent=sent, bytes_recv=recv)
    except Exception:
        return SimpleNamespace(bytes_sent=0, bytes_recv=0)


def _fallback_boot_time() -> float:
    try:
        uptime_secs = float(Path("/proc/uptime").read_text().split()[0])
        return time.time() - uptime_secs
    except Exception:
        return time.time()


def _system_stats() -> dict:
    global _net_prev, _cpu_primed
    now = time.time()
    if _system_micro_cache["data"] is not None and now - _system_micro_cache["ts"] < 1.5:
        return dict(_system_micro_cache["data"])

    if psutil is not None:
        if not _cpu_primed:
            psutil.cpu_percent(interval=None)
            _cpu_primed = True
        cpu_pct = psutil.cpu_percent(interval=0)
        mem = psutil.virtual_memory()
        disk = psutil.disk_usage("/")
        net = psutil.net_io_counters()
        boot_ts = psutil.boot_time()
    else:
        cpu_pct = _fallback_cpu_percent()
        mem = _fallback_virtual_memory()
        disk = _fallback_disk_usage("/")
        net = _fallback_net_io()
        boot_ts = _fallback_boot_time()
    now = time.time()
    net_speed = {"up_kbps": 0.0, "down_kbps": 0.0}
    if _net_prev["ts"]:
        dt = now - _net_prev["ts"]
        if dt > 0:
            net_speed["up_kbps"] = round((net.bytes_sent - _net_prev["sent"]) / dt / 1024, 1)
            net_speed["down_kbps"] = round((net.bytes_recv - _net_prev["recv"]) / dt / 1024, 1)
    _net_prev = {"ts": now, "sent": net.bytes_sent, "recv": net.bytes_recv}

    bot_mem = sum(_proc_info(n).get("memory_mb", 0) for n in PROCESSES)
    bot_cpu = sum(_proc_info(n).get("cpu", 0) for n in PROCESSES)

    boot = datetime.fromtimestamp(boot_ts, tz=timezone.utc)
    uptime_secs = (datetime.now(timezone.utc) - boot).total_seconds()

    payload = {
        "cpu": {"total": cpu_pct, "bot": round(bot_cpu, 1)},
        "ram": {
            "total_gb": round(mem.total / 1024**3, 1),
            "used_pct": mem.percent,
            "bot_mb": round(bot_mem, 1),
        },
        "disk": {
            "total_gb": round(disk.total / 1024**3, 1),
            "used_pct": disk.percent,
            "free_gb": round(disk.free / 1024**3, 1),
        },
        "network": net_speed,
        "uptime_secs": int(uptime_secs),
        "hostname": os.uname().nodename,
    }
    _system_micro_cache["ts"] = now
    _system_micro_cache["data"] = payload
    return payload


def _audit(action: str, detail: str = "", user: str | None = None) -> None:
    try:
        AUDIT_LOG.parent.mkdir(parents=True, exist_ok=True)
        who = user or session.get("username") or "system"
        line = json.dumps(
            {
                "ts": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "user": who,
                "action": action,
                "detail": detail,
            },
            ensure_ascii=False,
        )
        with AUDIT_LOG.open("a", encoding="utf-8") as fh:
            fh.write(line + "\n")
    except Exception:
        pass


def _parse_audit_log(limit: int = 100) -> list[dict]:
    if not AUDIT_LOG.exists():
        return []
    rows: list[dict] = []
    for line in AUDIT_LOG.read_text(errors="replace").splitlines()[-limit:]:
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return list(reversed(rows))


_MQ5_VERSION_RE = re.compile(r'#property\s+version\s+"([^"]+)"', re.I)
_MQ5_INPUT_RE = re.compile(
    r"^\s*input\s+(?:bool|int|double|string|color|datetime|ulong|long|float)\s+(\w+)",
    re.I | re.M,
)
_MQ5_MIN_CONFLUENCE_RE = re.compile(
    r"^\s*input\s+int\s+InpMinConfluence\s*=\s*(\d+)",
    re.I | re.M,
)


def _ensure_strategy_dirs() -> None:
    STRATEGY_DIR.mkdir(parents=True, exist_ok=True)
    STRATEGY_UPLOADS.mkdir(parents=True, exist_ok=True)


def _load_strategy_manifest() -> dict:
    _ensure_strategy_dirs()
    if STRATEGY_MANIFEST.exists():
        try:
            data = json.loads(STRATEGY_MANIFEST.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                data.setdefault("active_id", None)
                data.setdefault("history", [])
                return data
        except Exception:
            pass
    return {"active_id": None, "history": []}


def _save_strategy_manifest(data: dict) -> None:
    _ensure_strategy_dirs()
    STRATEGY_MANIFEST.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def _analyze_mq5(content: str) -> dict:
    version = None
    m = _MQ5_VERSION_RE.search(content)
    if m:
        version = m.group(1)
    inputs = _MQ5_INPUT_RE.findall(content)
    title = None
    for line in content.splitlines()[:20]:
        if "SignalBot" in line and ".mq5" in line:
            title = line.strip(" /|")
            break
    return {
        "version": version,
        "title": title,
        "input_count": len(inputs),
        "inputs_preview": inputs[:12],
    }


def _strategy_entry_by_id(manifest: dict, entry_id: str) -> dict | None:
    for row in manifest.get("history", []):
        if row.get("id") == entry_id:
            return row
    return None


def _apply_strategy_file(entry: dict, content: str, *, restart_engine: bool = True) -> dict:
    _ensure_strategy_dirs()
    stored = STRATEGY_UPLOADS / entry["stored_name"]
    if not stored.exists():
        raise FileNotFoundError("Uploaded strategy file missing on server")
    STRATEGY_ACTIVE.write_text(content, encoding="utf-8")
    STRATEGY_LEGACY.write_text(content, encoding="utf-8")
    env_updates: dict[str, str] = {"STRATEGY_MQ5_PATH": str(STRATEGY_ACTIVE)}
    min_conf = _MQ5_MIN_CONFLUENCE_RE.search(content)
    if min_conf:
        env_updates["MIN_SCORE"] = min_conf.group(1)
    _write_env(env_updates)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    entry["applied_at"] = ts
    manifest = _load_strategy_manifest()
    manifest["active_id"] = entry["id"]
    for idx, row in enumerate(manifest.get("history", [])):
        if row.get("id") == entry["id"]:
            manifest["history"][idx] = entry
            break
    manifest.setdefault("activation_log", []).append(
        {
            "entry_id": entry["id"],
            "applied_at": ts,
            "version": entry.get("version"),
            "original_name": entry.get("original_name"),
        }
    )
    manifest["activation_log"] = manifest["activation_log"][-120:]
    _save_strategy_manifest(manifest)
    _audit("strategy_apply", f"{entry.get('original_name')} ({entry.get('id')})")
    restart_ok = None
    restart_skipped = False
    restart_note = ""
    if restart_engine:
        restart_ok, restart_note, restart_skipped = _restart_signal_engine()
        if not restart_ok:
            return {
                "ok": True,
                "applied": True,
                "restart_engine": False,
                "restart_skipped": restart_skipped,
                "restart_error": restart_note,
                "entry": entry,
                "min_score_synced": env_updates.get("MIN_SCORE"),
            }
    return {
        "ok": True,
        "applied": True,
        "restart_engine": restart_ok,
        "restart_skipped": restart_skipped,
        "restart_note": restart_note if restart_skipped else None,
        "entry": entry,
        "min_score_synced": env_updates.get("MIN_SCORE"),
    }


def _empty_strategy_perf() -> dict:
    return {
        "signals": 0,
        "buy": 0,
        "sell": 0,
        "wins": 0,
        "losses": 0,
        "open": 0,
        "closed": 0,
        "win_rate": None,
        "tp1": 0,
        "tp2": 0,
        "avg_score": None,
        "delivery_sent": 0,
        "delivery_rate": None,
        "activations": 0,
        "active_hours": 0.0,
        "first_active_at": None,
        "last_active_at": None,
        "top_symbols": [],
    }


def _parse_audit_strategy_activations() -> list[dict]:
    events: list[dict] = []
    for row in _parse_audit_log(limit=1000):
        if row.get("action") != "strategy_apply":
            continue
        detail = str(row.get("detail") or "")
        m = re.search(r"\((\d{8}-\d{6})\)", detail)
        if not m:
            continue
        ts = str(row.get("ts") or "").strip()
        if not ts:
            continue
        events.append({"entry_id": m.group(1), "applied_at": ts})
    return events


def _strategy_activation_events(manifest: dict) -> list[dict]:
    logged = list(manifest.get("activation_log") or [])
    if logged:
        events = logged
    else:
        events = []
        seen: set[tuple[str, str]] = set()
        for ev in _parse_audit_strategy_activations():
            key = (ev["entry_id"], ev["applied_at"])
            if key in seen:
                continue
            seen.add(key)
            events.append(ev)
        for row in manifest.get("history", []):
            applied = row.get("applied_at")
            if not applied:
                continue
            key = (row["id"], applied)
            if key in seen:
                continue
            seen.add(key)
            events.append({"entry_id": row["id"], "applied_at": applied})
    events.sort(
        key=lambda ev: _parse_ts(str(ev.get("applied_at", "")).replace(" UTC", ""))
        or datetime.min
    )
    return events


def _strategy_activation_periods(manifest: dict) -> list[dict]:
    active_id = manifest.get("active_id")
    events = _strategy_activation_events(manifest)
    now = datetime.now()
    periods: list[dict] = []
    for i, ev in enumerate(events):
        start = _parse_ts(str(ev.get("applied_at", "")).replace(" UTC", ""))
        if not start:
            continue
        end = (
            _parse_ts(str(events[i + 1].get("applied_at", "")).replace(" UTC", ""))
            if i + 1 < len(events)
            else now
        )
        if not end:
            end = now
        duration_secs = max(0, int((end - start).total_seconds()))
        periods.append(
            {
                "entry_id": ev["entry_id"],
                "started_at": ev.get("applied_at"),
                "ended_at": events[i + 1].get("applied_at") if i + 1 < len(events) else None,
                "is_active": (i + 1 >= len(events)) and ev["entry_id"] == active_id,
                "duration_secs": duration_secs,
                "_start": start,
                "_end": end,
            }
        )
    return periods


def _signals_between(signals: list[dict], start: datetime, end: datetime) -> list[dict]:
    out: list[dict] = []
    for sig in signals:
        ts = _parse_ts(str(sig.get("timestamp", ""))[:19])
        if not ts or ts < start or ts >= end:
            continue
        out.append(sig)
    return out


def _compute_strategy_signal_stats(signals: list[dict]) -> dict:
    if not signals:
        return _empty_strategy_perf()
    wins = [s for s in signals if s.get("outcome") in ("tp1", "tp2")]
    losses = [s for s in signals if s.get("outcome") == "sl"]
    open_ = [s for s in signals if s.get("outcome") in (None, "open")]
    closed = len(wins) + len(losses)
    buy = sum(1 for s in signals if str(s.get("direction", "")).upper() == "BUY")
    sell = sum(1 for s in signals if str(s.get("direction", "")).upper() == "SELL")
    scores = [int(s["score"]) for s in signals if s.get("score") is not None]
    sent = [s for s in signals if s.get("delivery_status") == "sent"]
    sym_counts: dict[str, int] = defaultdict(int)
    for s in signals:
        sym = str(s.get("symbol") or "?")
        sym_counts[sym] += 1
    top_symbols = sorted(sym_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    return {
        "signals": len(signals),
        "buy": buy,
        "sell": sell,
        "wins": len(wins),
        "losses": len(losses),
        "open": len(open_),
        "closed": closed,
        "win_rate": round(len(wins) / closed * 100, 1) if closed else None,
        "tp1": sum(1 for s in signals if s.get("outcome") == "tp1"),
        "tp2": sum(1 for s in signals if s.get("outcome") == "tp2"),
        "avg_score": round(sum(scores) / len(scores), 1) if scores else None,
        "delivery_sent": len(sent),
        "delivery_rate": round(len(sent) / len(signals) * 100, 1) if signals else None,
        "top_symbols": [{"symbol": sym, "count": cnt} for sym, cnt in top_symbols],
    }


def _merge_strategy_perf(base: dict, extra: dict) -> dict:
    merged = dict(base)
    for key in ("signals", "buy", "sell", "wins", "losses", "open", "closed", "tp1", "tp2", "delivery_sent"):
        merged[key] = base.get(key, 0) + extra.get(key, 0)
    closed = merged["closed"]
    merged["win_rate"] = round(merged["wins"] / closed * 100, 1) if closed else None
    total = merged["signals"]
    merged["delivery_rate"] = round(merged["delivery_sent"] / total * 100, 1) if total else None
    sym_map: dict[str, int] = defaultdict(int)
    for row in base.get("top_symbols") or []:
        sym_map[row["symbol"]] += row["count"]
    for row in extra.get("top_symbols") or []:
        sym_map[row["symbol"]] += row["count"]
    merged["top_symbols"] = [
        {"symbol": sym, "count": cnt}
        for sym, cnt in sorted(sym_map.items(), key=lambda x: x[1], reverse=True)[:5]
    ]
    scores_weight = []
    if base.get("avg_score") is not None and base.get("signals"):
        scores_weight.append((base["avg_score"], base["signals"]))
    if extra.get("avg_score") is not None and extra.get("signals"):
        scores_weight.append((extra["avg_score"], extra["signals"]))
    if scores_weight:
        total_w = sum(w for _, w in scores_weight)
        merged["avg_score"] = round(sum(v * w for v, w in scores_weight) / total_w, 1)
    else:
        merged["avg_score"] = None
    return merged


def _strategy_performance_by_entry(
    manifest: dict,
    signals: list[dict],
    periods: list[dict],
) -> dict[str, dict]:
    perf: dict[str, dict] = {}
    meta_by_id = {row["id"]: row for row in manifest.get("history", []) if row.get("id")}
    grouped: dict[str, list[dict]] = defaultdict(list)
    for period in periods:
        grouped[period["entry_id"]].append(period)
    for entry_id, entry_periods in grouped.items():
        base = _empty_strategy_perf()
        base["activations"] = len(entry_periods)
        base["active_hours"] = round(sum(p["duration_secs"] for p in entry_periods) / 3600, 1)
        starts = [p.get("started_at") for p in entry_periods if p.get("started_at")]
        if starts:
            base["first_active_at"] = min(starts)
            base["last_active_at"] = max(starts)
        for period in entry_periods:
            chunk = _signals_between(signals, period["_start"], period["_end"])
            stats = _compute_strategy_signal_stats(chunk)
            base = _merge_strategy_perf(base, stats)
        if entry_periods and entry_periods[-1].get("is_active"):
            base["is_active_now"] = True
        entry = meta_by_id.get(entry_id) or {}
        base["version"] = entry.get("version")
        base["original_name"] = entry.get("original_name")
        perf[entry_id] = base
    return perf


def _strategy_legacy_stats(signals: list[dict], periods: list[dict]) -> dict:
    if not signals:
        return _empty_strategy_perf()
    if not periods:
        return _compute_strategy_signal_stats(signals)
    first_start = min(p["_start"] for p in periods)
    legacy = _signals_between(signals, datetime.min, first_start)
    stats = _compute_strategy_signal_stats(legacy)
    stats["label"] = "قبل از ثبت نسخه‌ها"
    return stats


def _strategy_status_payload() -> dict:
    manifest = _load_strategy_manifest()
    active = _strategy_entry_by_id(manifest, manifest.get("active_id") or "")
    active_exists = STRATEGY_ACTIVE.exists()
    legacy_exists = STRATEGY_LEGACY.exists()
    enriched = _get_enriched_all()
    periods_raw = _strategy_activation_periods(manifest)
    perf_by_id = _strategy_performance_by_entry(manifest, enriched, periods_raw)
    legacy_perf = _strategy_legacy_stats(enriched, periods_raw)
    periods_public = [
        {k: v for k, v in p.items() if not k.startswith("_")} for p in periods_raw
    ]
    history: list[dict] = []
    for row in manifest.get("history", [])[:20]:
        item = dict(row)
        item["performance"] = perf_by_id.get(row.get("id") or "", _empty_strategy_perf())
        history.append(item)
    active_id = manifest.get("active_id")
    active_perf = perf_by_id.get(active_id, _empty_strategy_perf()) if active_id else None
    comparison = []
    for row in history:
        perf = row.get("performance") or _empty_strategy_perf()
        if not perf.get("signals") and not row.get("applied_at"):
            continue
        comparison.append(
            {
                "id": row.get("id"),
                "original_name": row.get("original_name"),
                "version": row.get("version"),
                "is_active": row.get("id") == active_id,
                "applied_at": row.get("applied_at"),
                "win_rate": perf.get("win_rate"),
                "signals": perf.get("signals"),
                "wins": perf.get("wins"),
                "losses": perf.get("losses"),
                "closed": perf.get("closed"),
                "avg_score": perf.get("avg_score"),
                "active_hours": perf.get("active_hours"),
            }
        )
    comparison.sort(
        key=lambda r: _parse_ts(str(r.get("applied_at") or "").replace(" UTC", ""))
        or datetime.min,
        reverse=True,
    )
    rated = [c for c in comparison if c.get("win_rate") is not None and (c.get("closed") or 0) > 0]
    best = max(rated, key=lambda c: c["win_rate"], default=None)
    return {
        "active": active,
        "active_id": active_id,
        "active_path": str(STRATEGY_ACTIVE) if active_exists else None,
        "legacy_path": str(STRATEGY_LEGACY) if legacy_exists else None,
        "history": history,
        "uploads_dir": str(STRATEGY_UPLOADS),
        "performance_summary": {
            "active": active_perf,
            "legacy": legacy_perf,
            "best_win_rate_id": best.get("id") if best else None,
            "comparison": comparison,
            "activation_timeline": periods_public,
            "tracked_signals": len(enriched),
        },
    }


def _ops_config() -> dict:
    env = _parse_env()
    return {
        "alert_cpu_threshold": int(env.get("ALERT_CPU_THRESHOLD", "90") or 90),
        "alert_ram_threshold": int(env.get("ALERT_RAM_THRESHOLD", "90") or 90),
        "alert_disk_threshold": int(env.get("ALERT_DISK_THRESHOLD", "92") or 92),
        "webhook_discord_url": env.get("WEBHOOK_DISCORD_URL", ""),
        "webhook_slack_url": env.get("WEBHOOK_SLACK_URL", ""),
        "webhook_on_signal": env.get("WEBHOOK_ON_SIGNAL", "0") == "1",
        "maintenance_enabled": env.get("MAINTENANCE_ENABLED", "0") == "1",
        "maintenance_window": env.get("MAINTENANCE_WINDOW", "22:00-06:00"),
    }


def _record_metrics(stats: dict) -> None:
    METRICS_HISTORY.append(
        {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "cpu": stats.get("cpu", {}).get("total"),
            "ram": stats.get("ram", {}).get("used_pct"),
            "disk": stats.get("disk", {}).get("used_pct"),
            "net_down": stats.get("network", {}).get("down_kbps"),
            "net_up": stats.get("network", {}).get("up_kbps"),
        }
    )


def _check_resource_alerts(stats: dict) -> list[dict]:
    cfg = _ops_config()
    alerts: list[dict] = []
    checks = [
        ("cpu", stats.get("cpu", {}).get("total", 0), cfg["alert_cpu_threshold"], "CPU"),
        ("ram", stats.get("ram", {}).get("used_pct", 0), cfg["alert_ram_threshold"], "RAM"),
        ("disk", stats.get("disk", {}).get("used_pct", 0), cfg["alert_disk_threshold"], "Disk"),
    ]
    now = time.time()
    for key, value, threshold, label in checks:
        if value is None or threshold <= 0:
            continue
        if float(value) >= float(threshold):
            last = _LAST_ALERT_TS.get(key, 0)
            if now - last > 300:
                _LAST_ALERT_TS[key] = now
                msg = f"{label} usage {value}% (threshold {threshold}%)"
                alerts.append({"type": key, "message": msg, "value": value, "threshold": threshold})
                # Resource alerts stay dashboard-only; Telegram is reserved for trading signals.
    return alerts


def _pm2_restarts_24h() -> dict:
    cutoff = datetime.now() - timedelta(hours=24)
    counts: dict[str, int] = {name: 0 for name in DISPLAY_PROCESSES}
    events: dict[str, list[str]] = {name: [] for name in DISPLAY_PROCESSES}
    if PM2_EVENT_LOG.exists():
        for line in PM2_EVENT_LOG.read_text(errors="replace").splitlines():
            if "App [" not in line:
                continue
            m = re.search(r"App \[([\w-]+):\d+\]", line)
            if not m:
                continue
            name = m.group(1)
            if name not in counts:
                continue
            ts_match = re.match(r"(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})", line)
            if not ts_match:
                continue
            try:
                ts = datetime.strptime(ts_match.group(1), "%Y-%m-%dT%H:%M:%S")
            except ValueError:
                continue
            if ts < cutoff:
                continue
            if any(k in line.lower() for k in ("starting", "online", "exited with code", "stopped")):
                counts[name] += 1
                events[name].append(ts.strftime("%H:%M %d/%m"))
    procs = []
    for name in DISPLAY_PROCESSES:
        info = _proc_info(name)
        procs.append(
            {
                "name": name,
                "status": info.get("status"),
                "restarts_total": info.get("restarts", 0),
                "restarts_24h": counts.get(name, 0),
                "events_24h": events.get(name, [])[-8:],
                "uptime_human": _uptime_str(info.get("uptime")),
            }
        )
    return {"processes": procs, "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M")}


def _webhook_post(url: str, payload: dict, kind: str = "discord") -> bool:
    if not url:
        return False
    try:
        if kind == "slack":
            body = json.dumps({"text": payload.get("text", "")}).encode("utf-8")
        else:
            body = json.dumps({"content": payload.get("text", "")[:1900]}).encode("utf-8")
        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=10) as resp:
            return 200 <= resp.status < 300
    except Exception:
        return False


def _notify_signal_webhooks(sig: dict) -> None:
    cfg = _ops_config()
    if not cfg["webhook_on_signal"]:
        return
    sym = sig.get("symbol", "?")
    direction = str(sig.get("direction", "")).upper()
    text = f"New signal: {direction} {sym} entry={sig.get('entry', '—')}"
    if cfg["webhook_discord_url"]:
        _webhook_post(cfg["webhook_discord_url"], {"text": text}, "discord")
    if cfg["webhook_slack_url"]:
        _webhook_post(cfg["webhook_slack_url"], {"text": text}, "slack")


def _detect_new_signal(latest: dict | None) -> bool:
    global _LAST_SIGNAL_KEY
    if not latest:
        return False
    key = f"{latest.get('symbol')}:{latest.get('timestamp')}:{latest.get('direction')}"
    if key == _LAST_SIGNAL_KEY:
        return False
    if _LAST_SIGNAL_KEY is not None:
        _LAST_SIGNAL_KEY = key
        return True
    _LAST_SIGNAL_KEY = key
    return False


def _in_maintenance_window(window: str) -> bool:
    try:
        start_s, end_s = window.split("-", 1)
        now = datetime.now().time()
        start = datetime.strptime(start_s.strip(), "%H:%M").time()
        end = datetime.strptime(end_s.strip(), "%H:%M").time()
        if start <= end:
            return start <= now <= end
        return now >= start or now <= end
    except Exception:
        return False


def _apply_maintenance_schedule() -> None:
    cfg = _ops_config()
    if not cfg["maintenance_enabled"]:
        if _MAINTENANCE_STATE["paused_by_schedule"]:
            _MAINTENANCE_STATE["paused_by_schedule"] = False
        return
    env = _parse_env()
    in_window = _in_maintenance_window(cfg["maintenance_window"])
    paused = env.get("NOTIFICATIONS_PAUSED", "0") == "1"
    if in_window and not paused and not _MAINTENANCE_STATE["paused_by_schedule"]:
        _write_env({"NOTIFICATIONS_PAUSED": "1"})
        _run(["pm2", "stop", "signal-engine"])
        _MAINTENANCE_STATE["paused_by_schedule"] = True
        _audit("maintenance_auto_pause", cfg["maintenance_window"])
    elif not in_window and _MAINTENANCE_STATE["paused_by_schedule"]:
        _write_env({"NOTIFICATIONS_PAUSED": "0"})
        _run(["pm2", "start", "signal-engine"])
        _MAINTENANCE_STATE["paused_by_schedule"] = False
        _audit("maintenance_auto_resume", cfg["maintenance_window"])


def _telegram_reply(chat_id: str, text: str) -> dict:
    env = _parse_env()
    token = env.get("TELEGRAM_BOT_TOKEN", "").strip()
    if not token or not chat_id:
        return {"ok": False, "error": "not configured"}
    payload = json.dumps({"chat_id": chat_id, "text": text, "parse_mode": "HTML"}).encode("utf-8")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = json.loads(resp.read().decode("utf-8"))
            return {"ok": bool(body.get("ok")), "response": body}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def _handle_telegram_command(text: str, chat_id: str) -> None:
    cmd = (text or "").strip().split()[0].lower() if text else ""
    env = _parse_env()
    if cmd == "/status":
        procs = [_proc_info(n) for n in PROCESSES]
        overall = _overall_status(procs)
        stats = _live_signal_stats()
        msg = (
            f"📊 <b>TradeChi Status</b>\n"
            f"Overall: <b>{overall}</b>\n"
            f"Signals today: {stats.get('today', 0)}\n"
            f"Paused: {env.get('NOTIFICATIONS_PAUSED', '0') == '1'}\n"
            f"Symbols: {env.get('SYMBOLS', '—')}"
        )
        _telegram_reply(chat_id, msg)
    elif cmd == "/pause":
        _write_env({"NOTIFICATIONS_PAUSED": "1"})
        _run(["pm2", "stop", "signal-engine"])
        _audit("telegram_pause", "/pause command", user="telegram")
        _telegram_reply(chat_id, "⏸ Notifications paused")
    elif cmd == "/symbols":
        _telegram_reply(chat_id, f"📋 Symbols:\n<code>{env.get('SYMBOLS', '—')}</code>")
    elif cmd == "/resume":
        _write_env({"NOTIFICATIONS_PAUSED": "0"})
        _run(["pm2", "start", "signal-engine"])
        _audit("telegram_resume", "/resume command", user="telegram")
        _telegram_reply(chat_id, "▶ Notifications resumed")


def _system_stats_with_ops() -> dict:
    _apply_maintenance_schedule()
    stats = _system_stats()
    _record_metrics(stats)
    alerts = _check_resource_alerts(stats)
    stats["alerts"] = alerts
    stats["ops"] = _ops_config()
    return stats


def _overall_status(procs: list[dict]) -> str:
    statuses = [p["status"] for p in procs]
    if all(s == "online" for s in statuses):
        return "running"
    if all(s == "stopped" for s in statuses):
        return "stopped"
    if any(s == "online" for s in statuses):
        return "partial"
    return "unknown"


# ── Routes ────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "index.html")


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "")
    password = data.get("password", "")
    valid_user = secrets.compare_digest(username, _dashboard_username())
    valid_pass = secrets.compare_digest(password, _dashboard_password())
    if valid_user and valid_pass:
        session.permanent = True
        session["authenticated"] = True
        session["username"] = username
        _audit("login", f"user {username}", user=username)
        return jsonify({"ok": True, "username": username})
    return jsonify({"error": "Invalid username or password"}), 401


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/auth/check")
def auth_check():
    return jsonify({
        "authenticated": bool(session.get("authenticated")),
        "username": session.get("username"),
    })


def _build_status_payload(
    stats_signals: list[dict] | None = None,
    enriched_recent: list[dict] | None = None,
) -> dict:
    pm2_map = _pm2_by_name()
    procs = []
    for name in PROCESSES:
        info = _proc_info(name, pm2_map)
        info["uptime_human"] = _uptime_str(info.get("uptime"))
        info["controllable"] = True
        procs.append(info)

    all_procs = []
    for name in DISPLAY_PROCESSES:
        info = _proc_info(name, pm2_map)
        info["uptime_human"] = _uptime_str(info.get("uptime"))
        info["controllable"] = name in PROCESSES
        all_procs.append(info)

    state = _read_json(STATE_FILE) or {}
    env = _parse_env()
    signal_stats = _live_signal_stats()
    if enriched_recent is None:
        enriched_recent = _get_enriched_signals(days=7)

    last_signal_human = {}
    for sym, ts in (state.get("last_signal_at") or {}).items():
        try:
            last_signal_human[sym] = datetime.fromtimestamp(float(ts), tz=timezone.utc).strftime(
                "%Y-%m-%d %H:%M UTC"
            )
        except Exception:
            last_signal_human[sym] = "—"

    return {
        "overall": _overall_status(procs),
        "processes": procs,
        "all_processes": all_procs,
        "engine_state": {
            "last_bars": state.get("last_bars", {}),
            "last_signal_at": last_signal_human,
            "updated_at": state.get("updated_at"),
            "startup_sent": state.get("startup_sent", False),
        },
        "config": {
            "symbols": env.get("SYMBOLS", "EUR/USD,GBP/USD,XAU/USD"),
            "min_score": env.get("MIN_SCORE", "5"),
            "poll_seconds": env.get("POLL_SECONDS", "30"),
            "data_provider": env.get("DATA_PROVIDER", "twelvedata"),
            "facebook_enable": env.get("FACEBOOK_ENABLE", "1") == "1",
            "telegram_configured": bool(env.get("TELEGRAM_BOT_TOKEN")),
            "notifications_paused": env.get("NOTIFICATIONS_PAUSED", "0") == "1",
            "engine_debug": env.get("ENGINE_DEBUG", "0") == "1",
        },
        "latest_signal": _read_json(SIGNAL_QUEUE),
        "signal_stats": signal_stats,
        "recent_signals": enriched_recent[:5],
        "outcome_summary": _outcome_summary(enriched_recent),
        "delivery_summary": {
            "sent": len([s for s in enriched_recent if s.get("delivery_status") == "sent"]),
            "total": len(enriched_recent),
            "rate": round(
                len([s for s in enriched_recent if s.get("delivery_status") == "sent"])
                / max(len(enriched_recent), 1)
                * 100,
                1,
            )
            if enriched_recent
            else None,
        },
        "server_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _signal_after(s: dict, cutoff: datetime) -> bool:
    try:
        return datetime.strptime(s.get("timestamp", "")[:19], "%Y-%m-%d %H:%M:%S") >= cutoff
    except ValueError:
        return True


@app.route("/api/status")
@auth_required
def api_status():
    return jsonify(_build_status_payload())


@app.route("/api/version")
def api_version():
    return jsonify(_dashboard_version())


@app.route("/api/client-log", methods=["POST"])
@auth_required
def api_client_log():
    payload = request.get_json(silent=True) or {}
    line = {
        "time": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "ip": request.headers.get("X-Forwarded-For", request.remote_addr),
        "ua": request.headers.get("User-Agent", ""),
        "path": request.referrer or request.path,
        "event": payload.get("event"),
        "detail": payload.get("detail"),
    }
    try:
        CLIENT_DIAGNOSTIC_LOG.parent.mkdir(parents=True, exist_ok=True)
        with CLIENT_DIAGNOSTIC_LOG.open("a", encoding="utf-8") as f:
            f.write(json.dumps(line, ensure_ascii=False) + "\n")
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    return jsonify({"ok": True})


@app.route("/api/bootstrap")
@auth_required
def api_bootstrap():
    """Single payload for fast dashboard boot — parse signal log once."""
    def _load():
        enriched_30 = _get_enriched_signals(days=30)
        enriched_7d = _get_enriched_signals(days=7)
        signals_7d = _parse_all_signals(days=7)
        telegram_30 = _parse_telegram_deliveries(days=30, limit=3000)

        return {
            "version": _dashboard_version(),
            "status": _build_status_payload(enriched_recent=enriched_7d or enriched_30[:10]),
            "system": _system_stats_with_ops(),
            "signals": {
                "signals": enriched_30[:50],
                "summary": _signals_page_summary(enriched_30),
            },
            "report_7": _report_summary(signals_7d, 7),
            "telegram": {
                "summary": _telegram_summary(telegram_30, 30),
                "entries": telegram_30[:100],
            },
            "ops": _ops_config(),
            "cooldowns": _cooldown_status(),
            "uptime": _pm2_restarts_24h(),
        }

    return jsonify(_cache_json("bootstrap:v1", 10, _load))


@app.after_request
def _cache_headers(response):
    path = request.path or ""
    if path.startswith("/static/"):
        response.cache_control.no_store = True
        response.cache_control.no_cache = True
        response.cache_control.must_revalidate = True
        response.cache_control.max_age = 0
    elif path == "/" or path.endswith(".html"):
        response.cache_control.no_store = True
        response.cache_control.no_cache = True
        response.cache_control.must_revalidate = True
    elif path.startswith("/api/") and path not in ("/api/stream", "/api/auth/login", "/api/auth/logout", "/api/version"):
        response.cache_control.no_store = True
        response.cache_control.private = True
        response.cache_control.max_age = 0
        response.cache_control.must_revalidate = True
    return response


@app.route("/api/system")
@auth_required
def api_system():
    return jsonify(_system_stats_with_ops())


@app.route("/api/signals")
@auth_required
def api_signals():
    days = request.args.get("days", 30, type=int)
    limit = request.args.get("limit", 100, type=int)
    delivery = request.args.get("delivery", "all")
    direction = request.args.get("direction", "all")
    outcome = request.args.get("outcome", "all")
    symbol = request.args.get("symbol", "").strip().upper()

    days = min(max(days, 1), 365)
    limit = min(max(limit, 1), 500)

    cache_key = f"signals:v1:{days}:{limit}:{delivery}:{direction}:{outcome}:{symbol}"

    def _load():
        enriched = _get_enriched_signals(days=days)

        if delivery in ("sent", "failed", "unsent"):
            enriched = [s for s in enriched if s.get("delivery_status") == delivery]
        elif delivery == "mistaken":
            enriched = [s for s in enriched if s.get("quality") in ("mistaken", "duplicate")]
        elif delivery == "duplicate":
            enriched = [s for s in enriched if s.get("duplicate")]

        if direction in ("BUY", "SELL"):
            enriched = [s for s in enriched if str(s.get("direction", "")).upper() == direction]

        if outcome in ("tp1", "tp2", "sl", "open", "win", "loss"):
            if outcome == "win":
                enriched = [s for s in enriched if s.get("outcome") in ("tp1", "tp2")]
            elif outcome == "loss":
                enriched = [s for s in enriched if s.get("outcome") == "sl"]
            else:
                enriched = [s for s in enriched if s.get("outcome") == outcome]

        if symbol:
            enriched = [s for s in enriched if symbol in _normalize_symbol(str(s.get("symbol", ""))).replace("/", "")]

        summary = _signals_page_summary(enriched)
        return {
            "signals": enriched[:limit],
            "summary": summary,
            "stats": _signal_stats(enriched),
        }

    return jsonify(_cache_json(cache_key, 15, _load))


@app.route("/api/logs")
@auth_required
def api_logs():
    process = request.args.get("process", "signal-engine")
    lines = request.args.get("lines", 80, type=int)
    log_map = {
        "signal-engine": PM2_LOG_DIR / "signal-engine-error.log",
        "signal-server-out": PM2_LOG_DIR / "signal-server-out.log",
        "signal-server-err": PM2_LOG_DIR / "signal-server-error.log",
        "signal-server": PM2_LOG_DIR / "signal-server-out.log",
        "facebook": SIGNAL_LOG,
        "telegram": TELEGRAM_DELIVERY_LOG,
    }
    path = log_map.get(process, PM2_LOG_DIR / f"{process}-error.log")
    if not path.exists():
        return jsonify({"lines": [], "source": str(path)})
    content = path.read_text(errors="replace").splitlines()
    return jsonify({"lines": content[-lines:], "source": str(path)})


@app.route("/api/config")
@auth_required
def api_config_get():
    env = _parse_env()
    safe = {}
    for k, v in env.items():
        safe[k] = _mask(v) if k in SECRET_KEYS else v
    return jsonify({"config": safe})


@app.route("/api/config", methods=["PATCH"])
@auth_required
def api_config_patch():
    data = request.get_json(silent=True) or {}
    allowed = {
        "SYMBOLS",
        "MIN_SCORE",
        "POLL_SECONDS",
        "FACEBOOK_ENABLE",
        "DATA_PROVIDER",
        "ENGINE_DEBUG",
        "SEND_STARTUP_MESSAGE",
        "NOTIFICATIONS_PAUSED",
        *OPS_CONFIG_KEYS,
    }
    updates = {k: str(v) for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "No valid fields"}), 400
    _write_env(updates)
    _audit("config_update", ", ".join(f"{k}={updates[k]}" for k in updates.keys()))
    _invalidate_dashboard_cache()
    env = _parse_env()
    return jsonify({
        "ok": True,
        "updated": list(updates.keys()),
        "config": {k: env.get(k, "") for k in updates.keys()},
    })


@app.route("/api/control", methods=["POST"])
@auth_required
def api_control():
    data = request.get_json(silent=True) or {}
    action = data.get("action", "")
    target = data.get("process", "all")

    if action not in ("start", "stop", "restart"):
        return jsonify({"error": "Invalid action"}), 400

    targets = list(PROCESSES) if target == "all" else [target]
    if any(t not in PROCESSES for t in targets):
        return jsonify({"error": "Invalid process"}), 400

    results = []
    for t in targets:
        if t == "signal-engine" and action in ("start", "restart") and _notifications_paused():
            results.append(
                {
                    "process": t,
                    "ok": False,
                    "output": "نوتیفیکیشن متوقف است — از «ادامه نوتیفیکیشن» استفاده کنید",
                }
            )
            _stop_signal_engine()
            continue
        if action == "restart":
            code, out, err = _run(["pm2", "restart", t])
        else:
            code, out, err = _run(["pm2", action, t])
        results.append({"process": t, "ok": code == 0, "output": out or err})

    if action in ("start", "stop", "restart"):
        _run(["pm2", "save"])
        _invalidate_dashboard_cache()

    return jsonify({"results": results, "overall": _overall_status([_proc_info(t) for t in PROCESSES])})


@app.route("/api/reports/summary")
@auth_required
def api_reports_summary():
    days = request.args.get("days", 30, type=int)
    days = min(days, 365)
    cache_key = f"reports:summary:v1:{days}"

    def _load():
        signals = _parse_all_signals(days=days)
        summary = _report_summary(signals, days=days)
        env = _parse_env()
        summary["notifications_paused"] = env.get("NOTIFICATIONS_PAUSED", "0") == "1"
        summary["engine_debug"] = env.get("ENGINE_DEBUG", "0") == "1"
        return summary

    return jsonify(_cache_json(cache_key, 30, _load))


@app.route("/api/telegram/log")
@auth_required
def api_telegram_log():
    days = request.args.get("days", 30, type=int)
    limit = request.args.get("limit", 100, type=int)
    status = request.args.get("status", "all")
    cache_key = f"telegram:log:v1:{days}:{limit}:{status}"

    def _load():
        entries = _parse_telegram_deliveries(days=min(days, 365), limit=min(limit, 500))
        if status == "ok":
            entries = [e for e in entries if e.get("ok")]
        elif status == "failed":
            entries = [e for e in entries if not e.get("ok")]
        summary = _telegram_summary(_parse_telegram_deliveries(days=min(days, 365), limit=500), days=days)
        env = _parse_env()
        return {
            "entries": entries,
            "summary": summary,
            "telegram_configured": bool(env.get("TELEGRAM_BOT_TOKEN")),
            "notifications_paused": env.get("NOTIFICATIONS_PAUSED", "0") == "1",
        }

    return jsonify(_cache_json(cache_key, 30, _load))


@app.route("/api/telegram/summary")
@auth_required
def api_telegram_summary():
    days = request.args.get("days", 30, type=int)
    cache_key = f"telegram:summary:v1:{days}"

    def _load():
        entries = _parse_telegram_deliveries(days=min(days, 365), limit=500)
        summary = _telegram_summary(entries, days=days)
        env = _parse_env()
        summary["telegram_configured"] = bool(env.get("TELEGRAM_BOT_TOKEN"))
        summary["notifications_paused"] = env.get("NOTIFICATIONS_PAUSED", "0") == "1"
        return summary

    return jsonify(_cache_json(cache_key, 30, _load))


@app.route("/api/telegram/test", methods=["POST"])
@auth_required
def api_telegram_test():
    env = _parse_env()
    if not env.get("TELEGRAM_BOT_TOKEN"):
        return jsonify({"ok": False, "error": "تلگرام تنظیم نشده"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = (
        "✅ <b>TradeChi Bot — Test</b>\n"
        f"Dashboard test message\n"
        f"<code>{now}</code>"
    )
    result = _telegram_send(text)
    _log_telegram_delivery(
        symbol="TEST",
        direction="INFO",
        ok=result.get("ok", False),
        error=result.get("error"),
        http_status=result.get("http_status"),
        message_type="test",
    )
    _invalidate_dashboard_cache()
    if not result.get("ok"):
        return jsonify(result), 502
    return jsonify({"ok": True, "message": "پیام تست ارسال شد", **result})


@app.route("/api/telegram/retry", methods=["POST"])
@auth_required
def api_telegram_retry():
    data = request.get_json(silent=True) or {}
    symbol = str(data.get("symbol", "")).strip()
    if not symbol:
        return jsonify({"error": "symbol required"}), 400
    timestamp = str(data.get("timestamp", "")).strip()
    direction = str(data.get("direction", "")).strip()
    sig = _find_signal_for_retry(symbol, timestamp, direction)
    if not sig:
        sig = {
            "symbol": _normalize_symbol(symbol),
            "direction": direction.upper() or "—",
            "entry": data.get("entry"),
            "sl": data.get("sl"),
            "tp1": data.get("tp1"),
            "tp2": data.get("tp2"),
            "rr": data.get("rr"),
            "score": data.get("score"),
            "basis": data.get("basis") or "Retry from dashboard",
        }
    text = _format_signal_telegram_message(sig)
    result = _telegram_send(text)
    _log_telegram_delivery(
        symbol=sig.get("symbol", symbol),
        direction=str(sig.get("direction", direction or "—")),
        ok=result.get("ok", False),
        error=result.get("error"),
        score=sig.get("score"),
        entry=sig.get("entry"),
        http_status=result.get("http_status"),
        message_type="signal",
    )
    if not result.get("ok"):
        return jsonify({**result, "message": "ارسال مجدد ناموفق"}), 502
    return jsonify({"ok": True, "message": "سیگنال مجدداً ارسال شد", **result})


def _symbol_analytics_from(enriched: list[dict]) -> list[dict]:
    buckets: dict[str, dict] = {}
    for s in enriched:
        sym = _normalize_symbol(str(s.get("symbol", "?")))
        row = buckets.setdefault(
            sym,
            {
                "symbol": sym,
                "total": 0,
                "buy": 0,
                "sell": 0,
                "wins": 0,
                "losses": 0,
                "open": 0,
                "sent": 0,
                "failed": 0,
                "unsent": 0,
            },
        )
        row["total"] += 1
        direction = str(s.get("direction", "")).upper()
        if direction == "BUY":
            row["buy"] += 1
        elif direction == "SELL":
            row["sell"] += 1
        outcome = s.get("outcome", "open")
        if outcome in ("tp1", "tp2"):
            row["wins"] += 1
        elif outcome == "sl":
            row["losses"] += 1
        else:
            row["open"] += 1
        status = s.get("delivery_status", "unsent")
        if status in row:
            row[status] += 1
    result: list[dict] = []
    for row in buckets.values():
        closed = row["wins"] + row["losses"]
        row["win_rate"] = round(row["wins"] / closed * 100, 1) if closed else None
        row["delivery_rate"] = round(row["sent"] / row["total"] * 100, 1) if row["total"] else None
        result.append(row)
    result.sort(key=lambda r: r["total"], reverse=True)
    return result


def _symbol_analytics(days: int = 30) -> list[dict]:
    return _symbol_analytics_from(_get_enriched_signals(days=days))


def _cooldown_status() -> dict:
    state = _read_json(STATE_FILE) or {}
    env = _parse_env()
    try:
        cooldown_secs = int(env.get("SIGNAL_COOLDOWN_SECONDS", "14400"))
    except ValueError:
        cooldown_secs = 14400
    symbols_raw = env.get("SYMBOLS", "EUR/USD,GBP/USD,XAU/USD")
    symbols = [
        _normalize_symbol(part.strip())
        for part in symbols_raw.replace(";", ",").split(",")
        if part.strip()
    ]
    last_at_raw = state.get("last_signal_at") or {}
    now = datetime.now(timezone.utc)
    rows: list[dict] = []
    for sym in symbols:
        ts_val = None
        for key, val in last_at_raw.items():
            if _normalize_symbol(str(key)) == sym:
                ts_val = val
                break
        last_human = None
        age_secs = None
        ready = True
        remaining_secs = 0
        if ts_val is not None:
            try:
                ts = datetime.fromtimestamp(float(ts_val), tz=timezone.utc)
                last_human = ts.strftime("%Y-%m-%d %H:%M UTC")
                age_secs = int((now - ts).total_seconds())
                remaining_secs = max(0, cooldown_secs - age_secs)
                ready = remaining_secs <= 0
            except (TypeError, ValueError, OSError):
                pass
        rows.append(
            {
                "symbol": sym,
                "ready": ready,
                "last_signal_at": last_human,
                "age_seconds": age_secs,
                "remaining_seconds": remaining_secs,
                "cooldown_seconds": cooldown_secs,
            }
        )
    return {
        "cooldown_seconds": cooldown_secs,
        "symbols": rows,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def _hourly_distribution_from(signals: list[dict]) -> list[dict]:
    totals = [0] * 24
    buys = [0] * 24
    sells = [0] * 24
    for sig in signals:
        ts = _parse_ts(sig.get("timestamp", ""))
        if not ts:
            continue
        hour = ts.hour
        totals[hour] += 1
        direction = str(sig.get("direction", "")).upper()
        if direction == "BUY":
            buys[hour] += 1
        elif direction == "SELL":
            sells[hour] += 1
    return [
        {"hour": h, "label": f"{h:02d}:00", "total": totals[h], "buy": buys[h], "sell": sells[h]}
        for h in range(24)
    ]


def _hourly_distribution(days: int = 30) -> list[dict]:
    return _hourly_distribution_from(_parse_all_signals(days=days))


@app.route("/api/analytics/symbols")
@auth_required
def api_analytics_symbols():
    days = min(max(request.args.get("days", 30, type=int), 1), 365)
    cache_key = f"analytics:symbols:v1:{days}"

    def _load():
        return {
            "symbols": _symbol_analytics(days),
            "days": days,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }

    return jsonify(_cache_json(cache_key, 45, _load))


@app.route("/api/analytics/cooldowns")
@auth_required
def api_analytics_cooldowns():
    return jsonify(_cooldown_status())


@app.route("/api/analytics/hourly")
@auth_required
def api_analytics_hourly():
    days = min(max(request.args.get("days", 30, type=int), 1), 365)
    cache_key = f"analytics:hourly:v1:{days}"

    def _load():
        return {
            "hours": _hourly_distribution(days),
            "days": days,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }

    return jsonify(_cache_json(cache_key, 45, _load))


@app.route("/api/export/telegram.csv")
@auth_required
def export_telegram_csv():
    days = request.args.get("days", 30, type=int)
    entries = _parse_telegram_deliveries(days=min(days, 365), limit=2000)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Timestamp", "Symbol", "Direction", "Status", "Score", "Entry", "Error", "Detail", "Source"])
    for e in entries:
        writer.writerow([
            e.get("timestamp", ""),
            e.get("symbol", ""),
            e.get("direction", ""),
            e.get("status", ""),
            e.get("score", ""),
            e.get("entry", ""),
            e.get("error", ""),
            e.get("detail", ""),
            e.get("source", ""),
        ])
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    fname = f"telegram-log-{datetime.now().strftime('%Y%m%d')}.csv"
    return send_file(out, as_attachment=True, download_name=fname, mimetype="text/csv")


@app.route("/api/export/signals.xlsx")
@auth_required
def export_signals_xlsx():
    days = request.args.get("days", 30, type=int)
    signals = _parse_all_signals(days=min(days, 365))
    summary = _report_summary(signals, days=min(days, 365))
    if not HAS_OPENPYXL:
        return jsonify({"error": "openpyxl not installed"}), 500
    buf = _build_xlsx(signals, summary)
    fname = f"bot-report-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/export/signals.csv")
@auth_required
def export_signals_csv():
    days = request.args.get("days", 30, type=int)
    signals = _parse_all_signals(days=min(days, 365))
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Timestamp", "Symbol", "Direction", "Entry", "SL", "TP1", "TP2", "RR", "Basis"])
    for s in signals:
        writer.writerow([
            s.get("timestamp", ""), s.get("symbol", ""), s.get("direction", ""),
            s.get("entry", ""), s.get("sl", ""), s.get("tp1", ""), s.get("tp2", ""),
            s.get("rr", ""), s.get("basis", ""),
        ])
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    fname = f"signals-{datetime.now().strftime('%Y%m%d')}.csv"
    return send_file(out, as_attachment=True, download_name=fname, mimetype="text/csv")


@app.route("/api/management", methods=["POST"])
@auth_required
def api_management():
    data = request.get_json(silent=True) or {}
    action = data.get("action", "")

    if action == "restart_all":
        results = []
        for t in PROCESSES:
            if t == "signal-engine" and _notifications_paused():
                _stop_signal_engine()
                results.append(
                    {
                        "process": t,
                        "ok": True,
                        "skipped": True,
                        "message": "left stopped (notifications paused)",
                    }
                )
                continue
            code, out, err = _run(["pm2", "restart", t])
            results.append({"process": t, "ok": code == 0})
        _run(["pm2", "save"])
        _invalidate_dashboard_cache()
        return jsonify({"ok": True, "message": "All bot processes restarted", "results": results})

    if action == "restart_dashboard":
        code, out, err = _run(["pm2", "restart", "dashboard"])
        _run(["pm2", "save"])
        _invalidate_dashboard_cache()
        return jsonify({"ok": code == 0, "message": "Dashboard restarted", "output": out or err})

    if action == "flush_logs":
        code, out, err = _run(["pm2", "flush"])
        return jsonify({"ok": code == 0, "message": "PM2 logs cleared"})

    if action == "reset_cooldowns":
        state = _read_json(STATE_FILE) or {}
        state["last_signal_at"] = {}
        state["updated_at"] = datetime.now(timezone.utc).isoformat()
        STATE_FILE.write_text(json.dumps(state, indent=2, default=str))
        _invalidate_dashboard_cache()
        return jsonify({"ok": True, "message": "Signal cooldowns reset"})

    if action == "reset_startup_flag":
        state = _read_json(STATE_FILE) or {}
        state["startup_sent"] = False
        state["updated_at"] = datetime.now(timezone.utc).isoformat()
        STATE_FILE.write_text(json.dumps(state, indent=2, default=str))
        _invalidate_dashboard_cache()
        return jsonify({"ok": True, "message": "Startup message flag reset"})

    if action == "pause_notifications":
        _write_env({"NOTIFICATIONS_PAUSED": "1"})
        code, out, err = _stop_signal_engine()
        _run(["pm2", "save"])
        _invalidate_dashboard_cache()
        if code != 0:
            return jsonify({"ok": False, "error": f"توقف موتور ناموفق: {err or out or 'pm2 error'}"}), 500
        _audit("pause_notifications", "manual pause")
        return jsonify(
            {
                "ok": True,
                "message": "نوتیفیکیشن متوقف شد — موتور سیگنال خاموش شد و تا Resume ارسال نمی‌شود",
            }
        )

    if action == "resume_notifications":
        _write_env({"NOTIFICATIONS_PAUSED": "0"})
        ok, msg = _start_signal_engine(force=True)
        _run(["pm2", "save"])
        _invalidate_dashboard_cache()
        if not ok:
            return jsonify({"ok": False, "error": f"راه‌اندازی موتور ناموفق: {msg}"}), 500
        _audit("resume_notifications", "manual resume")
        return jsonify({"ok": True, "message": "نوتیفیکیشن از سر گرفته شد — موتور سیگنال روشن شد"})

    if action == "toggle_debug":
        env = _parse_env()
        current = env.get("ENGINE_DEBUG", "0") == "1"
        _write_env({"ENGINE_DEBUG": "0" if current else "1"})
        _invalidate_dashboard_cache()
        return jsonify({"ok": True, "debug": not current, "message": f"Debug mode {'enabled' if not current else 'disabled'}"})

    if action == "pull_and_restart":
        branch = data.get("branch") or os.environ.get("DEPLOY_BRANCH", "main")
        return jsonify(_git_deploy(branch))

    if action == "install_ssh_key":
        pubkey = data.get("public_key", "").strip()
        if not pubkey or not pubkey.startswith(("ssh-ed25519 ", "ssh-rsa ")):
            return jsonify({"error": "Invalid public_key"}), 400
        auth_file = Path("/root/.ssh/authorized_keys")
        auth_file.parent.mkdir(mode=0o700, exist_ok=True)
        existing = auth_file.read_text() if auth_file.exists() else ""
        if pubkey not in existing:
            auth_file.write_text(existing.rstrip() + "\n" + pubkey + "\n")
            auth_file.chmod(0o600)
        return jsonify({"ok": True, "message": "Deploy SSH key installed"})

    return jsonify({"error": f"Unknown action: {action}"}), 400


def _git_deploy(branch: str) -> dict:
    git_dir = BOT_ROOT
    if not (git_dir / ".git").exists():
        return {"ok": False, "error": "Not a git repository"}
    token = os.environ.get("GH_PAT") or os.environ.get("GITHUB_TOKEN", "")
    env = os.environ.copy()
    env["GIT_TERMINAL_PROMPT"] = "0"
    git_prefix = ["git"]
    if token:
        git_prefix = [
            "git",
            "-c",
            "credential.helper=!f() { echo username=sedshahab0; echo password=${GITHUB_TOKEN}; }; f",
        ]
        env["GITHUB_TOKEN"] = token
    steps = [
        git_prefix + ["-C", str(git_dir), "fetch", "origin", branch],
        git_prefix + ["-C", str(git_dir), "checkout", branch],
        git_prefix + ["-C", str(git_dir), "reset", "--hard", f"origin/{branch}"],
    ]
    outputs = []
    for cmd in steps:
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=120, env=env)
            code, out, err = r.returncode, r.stdout.strip(), r.stderr.strip()
        except subprocess.TimeoutExpired:
            code, out, err = 1, "", "timeout"
        except FileNotFoundError:
            code, out, err = 1, "", "command not found"
        outputs.append({"cmd": " ".join(cmd), "ok": code == 0, "output": out or err})
        if code != 0:
            return {"ok": False, "message": "Git pull failed", "steps": outputs}
    _run(["pm2", "restart", "dashboard"])
    _run(["pm2", "save"])
    _invalidate_dashboard_cache()
    return {"ok": True, "message": f"Pulled {branch} and restarted dashboard", "steps": outputs}


@app.route("/api/deploy/hook", methods=["POST"])
def api_deploy_hook():
    token = request.headers.get("X-Deploy-Token") or request.args.get("token", "")
    expected = os.environ.get("DEPLOY_HOOK_TOKEN", "")
    if not expected or not secrets.compare_digest(token, expected):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json(silent=True) or {}
    branch = data.get("branch") or os.environ.get("DEPLOY_BRANCH", "main")
    result = _git_deploy(branch)
    if not result.get("ok"):
        return jsonify(result), 500
    return jsonify(result)


@app.route("/api/health")
def api_health():
    procs = [_proc_info(n) for n in PROCESSES]
    overall = _overall_status(procs)
    version = _dashboard_version()
    body = {
        "status": overall,
        "healthy": overall in ("running", "partial"),
        "version": version.get("full"),
        "time": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    code = 200 if body["healthy"] else 503
    return jsonify(body), code


@app.route("/api/changelog")
def api_changelog():
    version = _dashboard_version()
    return jsonify({"current": version, "history": version.get("history", [])})


@app.route("/api/ops/config")
@auth_required
def api_ops_config_get():
    return jsonify(_ops_config())


@app.route("/api/ops/config", methods=["PATCH"])
@auth_required
def api_ops_config_patch():
    data = request.get_json(silent=True) or {}
    mapping = {
        "alert_cpu_threshold": "ALERT_CPU_THRESHOLD",
        "alert_ram_threshold": "ALERT_RAM_THRESHOLD",
        "alert_disk_threshold": "ALERT_DISK_THRESHOLD",
        "webhook_discord_url": "WEBHOOK_DISCORD_URL",
        "webhook_slack_url": "WEBHOOK_SLACK_URL",
        "webhook_on_signal": "WEBHOOK_ON_SIGNAL",
        "maintenance_enabled": "MAINTENANCE_ENABLED",
        "maintenance_window": "MAINTENANCE_WINDOW",
    }
    updates: dict[str, str] = {}
    for js_key, env_key in mapping.items():
        if js_key in data:
            val = data[js_key]
            if js_key in ("webhook_on_signal", "maintenance_enabled"):
                updates[env_key] = "1" if val else "0"
            else:
                updates[env_key] = str(val)
    if not updates:
        return jsonify({"error": "No valid fields"}), 400
    _write_env(updates)
    _audit("ops_config", ", ".join(updates.keys()))
    _invalidate_dashboard_cache()
    return jsonify({"ok": True, "config": _ops_config()})


@app.route("/api/ops/uptime")
@auth_required
def api_ops_uptime():
    return jsonify(_pm2_restarts_24h())


@app.route("/api/strategy")
@auth_required
def api_strategy_get():
    return jsonify(_strategy_status_payload())


@app.route("/api/strategy/performance")
@auth_required
def api_strategy_performance():
    entry_id = (request.args.get("id") or "").strip()
    manifest = _load_strategy_manifest()
    if not entry_id:
        entry_id = manifest.get("active_id") or ""
    entry = _strategy_entry_by_id(manifest, entry_id)
    if not entry:
        return jsonify({"error": "نسخه استراتژی یافت نشد"}), 404
    enriched = _get_enriched_all()
    periods_raw = _strategy_activation_periods(manifest)
    entry_periods = [p for p in periods_raw if p["entry_id"] == entry_id]
    perf = _strategy_performance_by_entry(manifest, enriched, periods_raw).get(
        entry_id, _empty_strategy_perf()
    )
    period_rows = []
    for period in entry_periods:
        chunk = _signals_between(enriched, period["_start"], period["_end"])
        stats = _compute_strategy_signal_stats(chunk)
        period_rows.append(
            {
                "started_at": period.get("started_at"),
                "ended_at": period.get("ended_at"),
                "is_active": period.get("is_active"),
                "duration_hours": round(period.get("duration_secs", 0) / 3600, 1),
                "stats": stats,
            }
        )
    version_signals: list[dict] = []
    for period in entry_periods:
        version_signals.extend(_signals_between(enriched, period["_start"], period["_end"]))
    version_signals.sort(key=lambda s: s.get("timestamp", ""), reverse=True)
    recent = []
    for sig in version_signals[:20]:
        recent.append(
            {
                "timestamp": sig.get("timestamp"),
                "symbol": sig.get("symbol"),
                "direction": sig.get("direction"),
                "score": sig.get("score"),
                "outcome": sig.get("outcome"),
                "delivery_status": sig.get("delivery_status"),
            }
        )
    return jsonify(
        {
            "entry": entry,
            "performance": perf,
            "periods": period_rows,
            "recent_signals": recent,
        }
    )


@app.route("/api/strategy/upload", methods=["POST"])
@auth_required
def api_strategy_upload():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"error": "فایلی انتخاب نشده است"}), 400
    original = secure_filename(upload.filename)
    if not original.lower().endswith(".mq5"):
        return jsonify({"error": "فقط فایل‌های .mq5 مجاز هستند"}), 400
    raw = upload.read()
    if not raw:
        return jsonify({"error": "فایل خالی است"}), 400
    if len(raw) > STRATEGY_MAX_BYTES:
        return jsonify({"error": "حداکثر اندازه فایل ۲ مگابایت است"}), 400
    try:
        content = raw.decode("utf-8")
    except UnicodeDecodeError:
        return jsonify({"error": "فایل باید متن UTF-8 باشد"}), 400
    if "input " not in content and "#property" not in content:
        return jsonify({"error": "فایل MQL5 معتبر به نظر نمی‌رسد"}), 400

    meta = _analyze_mq5(content)
    entry_id = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    stored_name = f"{entry_id}_{original}"
    _ensure_strategy_dirs()
    stored_path = STRATEGY_UPLOADS / stored_name
    stored_path.write_text(content, encoding="utf-8")
    checksum = hashlib.sha256(raw).hexdigest()
    who = session.get("username") or "admin"
    entry = {
        "id": entry_id,
        "original_name": original,
        "stored_name": stored_name,
        "uploaded_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "uploaded_by": who,
        "size_bytes": len(raw),
        "checksum_sha256": checksum,
        "version": meta.get("version"),
        "title": meta.get("title"),
        "input_count": meta.get("input_count"),
        "inputs_preview": meta.get("inputs_preview"),
        "applied_at": None,
    }
    manifest = _load_strategy_manifest()
    history = manifest.get("history", [])
    history.insert(0, entry)
    manifest["history"] = history[:30]
    _save_strategy_manifest(manifest)
    _audit("strategy_upload", f"{original} ({entry_id})")
    _invalidate_dashboard_cache()
    return jsonify({"ok": True, "entry": entry, "strategy": _strategy_status_payload()})


@app.route("/api/strategy/apply", methods=["POST"])
@auth_required
def api_strategy_apply():
    data = request.get_json(silent=True) or {}
    entry_id = (data.get("id") or "").strip()
    restart_engine = bool(data.get("restart_engine", True))
    manifest = _load_strategy_manifest()
    if not entry_id:
        entry_id = manifest.get("active_id") or ""
        if not entry_id and manifest.get("history"):
            entry_id = manifest["history"][0]["id"]
    entry = _strategy_entry_by_id(manifest, entry_id)
    if not entry:
        return jsonify({"error": "نسخه استراتژی یافت نشد"}), 404
    stored = STRATEGY_UPLOADS / entry["stored_name"]
    if not stored.exists():
        return jsonify({"error": "فایل آپلودشده روی سرور موجود نیست"}), 404
    content = stored.read_text(encoding="utf-8")
    try:
        result = _apply_strategy_file(entry, content, restart_engine=restart_engine)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    result["strategy"] = _strategy_status_payload()
    _invalidate_dashboard_cache()
    return jsonify(result)


@app.route("/api/strategy/download")
@auth_required
def api_strategy_download():
    entry_id = request.args.get("id", "").strip()
    use_active = request.args.get("active", "").lower() in ("1", "true", "yes")
    path: Path | None = None
    download_name = "strategy.mq5"
    if use_active or not entry_id:
        if STRATEGY_ACTIVE.exists():
            path = STRATEGY_ACTIVE
            download_name = "active.mq5"
        elif STRATEGY_LEGACY.exists():
            path = STRATEGY_LEGACY
            download_name = STRATEGY_LEGACY.name
    else:
        manifest = _load_strategy_manifest()
        entry = _strategy_entry_by_id(manifest, entry_id)
        if not entry:
            return jsonify({"error": "نسخه یافت نشد"}), 404
        path = STRATEGY_UPLOADS / entry["stored_name"]
        download_name = entry.get("original_name") or entry["stored_name"]
    if not path or not path.exists():
        return jsonify({"error": "فایل استراتژی موجود نیست"}), 404
    return send_file(path, as_attachment=True, download_name=download_name, mimetype="text/plain")


@app.route("/api/audit")
@auth_required
def api_audit():
    limit = min(max(request.args.get("limit", 100, type=int), 1), 500)
    return jsonify({"entries": _parse_audit_log(limit)})


@app.route("/api/config/backup")
@auth_required
def api_config_backup():
    env = _parse_env()
    export = {k: env[k] for k in sorted(BACKUP_KEYS) if k in env}
    for sk in SECRET_KEYS:
        export.pop(sk, None)
    _audit("config_backup", f"{len(export)} keys")
    return jsonify(
        {
            "exported_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "version": _dashboard_version().get("full"),
            "config": export,
        }
    )


@app.route("/api/config/restore", methods=["POST"])
@auth_required
def api_config_restore():
    data = request.get_json(silent=True) or {}
    cfg = data.get("config") if isinstance(data.get("config"), dict) else data
    updates = {k: str(v) for k, v in cfg.items() if k in BACKUP_KEYS and k not in SECRET_KEYS}
    if not updates:
        return jsonify({"error": "No valid config keys"}), 400
    _write_env(updates)
    _audit("config_restore", ", ".join(updates.keys()))
    _invalidate_dashboard_cache()
    return jsonify({"ok": True, "restored": list(updates.keys())})


@app.route("/api/export/metrics.csv")
@auth_required
def export_metrics_csv():
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Timestamp", "CPU %", "RAM %", "Disk %", "Net Down KB/s", "Net Up KB/s"])
    for row in METRICS_HISTORY:
        writer.writerow([
            row.get("ts", ""),
            row.get("cpu", ""),
            row.get("ram", ""),
            row.get("disk", ""),
            row.get("net_down", ""),
            row.get("net_up", ""),
        ])
    out = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    fname = f"metrics-{datetime.now().strftime('%Y%m%d-%H%M')}.csv"
    return send_file(out, as_attachment=True, download_name=fname, mimetype="text/csv")


@app.route("/api/telegram/webhook", methods=["POST"])
def api_telegram_webhook():
    data = request.get_json(silent=True) or {}
    message = data.get("message") or data.get("edited_message") or {}
    text = message.get("text") or ""
    chat = message.get("chat") or {}
    chat_id = str(chat.get("id", ""))
    env = _parse_env()
    allowed_chat = env.get("TELEGRAM_CHAT_ID", "").strip()
    if allowed_chat and chat_id and chat_id != allowed_chat:
        return jsonify({"ok": True, "ignored": True})
    if text.startswith("/"):
        _handle_telegram_command(text, chat_id or allowed_chat)
    return jsonify({"ok": True})


@app.route("/api/stream")
@auth_required
def api_stream():
    def generate():
        while True:
            procs = []
            all_procs = []
            for name in PROCESSES:
                info = _proc_info(name)
                info["uptime_human"] = _uptime_str(info.get("uptime"))
                info["controllable"] = True
                procs.append(info)
            for name in DISPLAY_PROCESSES:
                info = _proc_info(name)
                info["uptime_human"] = _uptime_str(info.get("uptime"))
                info["controllable"] = name in PROCESSES
                all_procs.append(info)
            latest = _read_json(SIGNAL_QUEUE)
            if _detect_new_signal(latest if isinstance(latest, dict) else None) and isinstance(latest, dict):
                _notify_signal_webhooks(latest)
            payload = {
                "overall": _overall_status(procs),
                "processes": procs,
                "all_processes": all_procs,
                "system": _system_stats_with_ops(),
                "signal_stats": _live_signal_stats(),
                "latest_signal": latest,
                "uptime_history": _pm2_restarts_24h(),
                "server_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            yield f"data: {json.dumps(payload)}\n\n"
            time.sleep(3)

    return Response(generate(), mimetype="text/event-stream", headers={"Cache-Control": "no-cache"})


if __name__ == "__main__":
    port = int(os.environ.get("DASHBOARD_PORT", "8080"))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
