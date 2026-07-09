"""
SCRIPT 2 — Facebook Group Auto Poster (Signal-Driven)
======================================================
Reads signal_queue.json (written by signal_server.py when MT4/MT5 fires a signal),
injects live values into templates, then posts to all groups in fb_my_groups.xlsx.

Can also be run manually — it will use whatever signal is currently in signal_queue.json.

REQUIREMENTS:
    pip install selenium openpyxl webdriver-manager

USAGE (manual):
    python script2_post_to_groups.py

USAGE (automatic):
    Triggered by signal_server.py when EA fires a signal.
"""

import argparse
import fcntl
import importlib.util
import sys
import time
import random
import json
import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
#  FILES
# ─────────────────────────────────────────────
GROUPS_FILE  = os.environ.get("FACEBOOK_GROUPS_FILE", "fb_my_groups.xlsx")
LOG_FILE     = os.environ.get("FACEBOOK_POST_LOG", "post_log.xlsx")
QUEUE_FILE   = os.environ.get("SIGNAL_QUEUE_FILE", "signal_queue.json")
SESSION_FILE = os.environ.get("FACEBOOK_SESSION_FILE", "fb_session.json")
LOCK_FILE    = os.environ.get("FACEBOOK_POSTER_LOCK", "/var/lib/trading-bot/facebook-poster.lock")
HEADLESS     = os.environ.get("FACEBOOK_HEADLESS", "1") == "1"

# ─────────────────────────────────────────────
#  SAFETY SETTINGS
# ─────────────────────────────────────────────
BATCH_SIZE        = 10
MIN_DELAY_SECONDS = 180
MAX_DELAY_SECONDS = 320

# ─────────────────────────────────────────────
#  SIGNAL LOADER
# ─────────────────────────────────────────────

def load_signal(signal_file=None):
    """
    Load signal from signal_queue.json.
    Falls back to a placeholder if file doesn't exist.
    """
    source = signal_file or QUEUE_FILE
    if os.path.exists(source):
        with open(source, "r", encoding="utf-8") as f:
            sig = json.load(f)
        print(f"[✓] Signal loaded: {sig['symbol']} {sig['direction']} @ {sig['entry']}")
        return sig
    raise FileNotFoundError(f"Signal file not found: {source}")


# ─────────────────────────────────────────────
#  TEMPLATE BUILDER  (dynamic — uses live signal)
# ─────────────────────────────────────────────

def symbol_label(symbol, lang):
    """Return a human-readable symbol label per language."""
    labels = {
        "XAUUSD": {"English": "Gold", "Persian": "طلا", "Russian": "Золото"},
        "EURUSD": {"English": "Euro/Dollar", "Persian": "یورو/دلار", "Russian": "Евро/Доллар"},
        "GBPUSD": {"English": "Cable", "Persian": "پوند/دلار", "Russian": "Фунт/Доллар"},
        "USDJPY": {"English": "Dollar/Yen", "Persian": "دلار/ین", "Russian": "Доллар/Иена"},
        "USDCHF": {"English": "Dollar/Franc", "Persian": "دلار/فرانک", "Russian": "Доллар/Франк"},
        "AUDUSD": {"English": "Aussie", "Persian": "دلار استرالیا", "Russian": "Австралийский доллар"},
        "USDCAD": {"English": "Loonie", "Persian": "دلار کانادا", "Russian": "Канадский доллар"},
        "SILVER": {"English": "Silver", "Persian": "نقره", "Russian": "Серебро"},
        "XAGUSD":{"English": "Silver", "Persian": "نقره", "Russian": "Серебро"},
    }
    return labels.get(symbol.upper(), {}).get(lang, symbol)


def direction_label(direction, lang):
    labels = {
        "BUY":  {"English": "BUY 📈",  "Persian": "خرید (BUY) 📈", "Russian": "ПОКУПКА (BUY) 📈"},
        "SELL": {"English": "SELL 📉", "Persian": "فروش (SELL) 📉", "Russian": "ПРОДАЖА (SELL) 📉"},
    }
    return labels.get(direction.upper(), {}).get(lang, direction)


def build_templates(sig):
    """
    Build all 9 templates dynamically using live signal values.
    Templates use placeholders that are replaced with actual signal data.
    """
    sym   = sig["symbol"]
    dirn  = sig["direction"].upper()
    entry = sig["entry"]
    sl    = sig["sl"]
    tp1   = sig["tp1"]
    tp2   = sig.get("tp2", "—")
    tp3   = sig.get("tp3", "—")
    rr    = sig.get("rr", "1:3")
    basis = sig.get("basis", "SMC Structure + Liquidity Grab + Daily Bias")

    # Build TP lines — skip if "—"
    def tp_lines_en(tp1, tp2, tp3):
        lines = f"✅ TP1: {tp1}\n"
        if tp2 != "—": lines += f"✅ TP2: {tp2}\n"
        if tp3 != "—": lines += f"✅ TP3: {tp3}\n"
        return lines

    def tp_lines_fa(tp1, tp2, tp3):
        lines = f"✅ TP1: {tp1}\n"
        if tp2 != "—": lines += f"✅ TP2: {tp2}\n"
        if tp3 != "—": lines += f"✅ TP3: {tp3}\n"
        return lines

    def tp_compact_en(tp2, tp3):
        if tp2 != "—" and tp3 != "—":
            return f"{tp2} / {tp3}"
        elif tp2 != "—":
            return tp2
        return tp1

    templates = {
        # ── ENGLISH ──────────────────────────────────────────────────────
        "English": {
            "1": (
                f"🟢 LIVE SIGNAL ALERT | {sym} ({symbol_label(sym, 'English')})\n\n"
                f"📈 Direction: {direction_label(dirn, 'English')}\n"
                f"🎯 Entry: {entry}\n"
                f"🛑 Stop Loss: {sl}\n"
                f"{tp_lines_en(tp1, tp2, tp3)}\n"
                f"⚡ Risk: 1% per trade\n"
                f"📊 Basis: {basis}\n\n"
                f"💬 DM me if you want to receive signals like this daily — FREE for a limited time.\n\n"
                f"#{sym} #ForexSignals #SMC #ICT #PriceAction #Trading"
            ),
            "2": (
                f"This setup is too clean to ignore.\n\n"
                f"🟢 {sym} — {'BUY' if dirn == 'BUY' else 'SELL'} OPPORTUNITY\n\n"
                f"{'Price swept sell-side liquidity and reacted at a premium order block.' if dirn == 'BUY' else 'Price swept buy-side liquidity and reacted at a discount order block.'} "
                f"Daily bias is {'bullish' if dirn == 'BUY' else 'bearish'}.\n\n"
                f"📍 Entry: {entry}\n"
                f"🛑 SL: {sl}\n"
                f"🎯 TP: {tp_compact_en(tp2, tp3)}\n\n"
                f"RR: {rr}\n\n"
                f"Want my signals before they go public? DM me.\n\n"
                f"#{sym} #ForexSignals #SMCTrading #ICT #PriceAction"
            ),
            "3": (
                f"Members in my Telegram channel get signals like this BEFORE I post publicly.\n\n"
                f"🟢 {sym} — {direction_label(dirn, 'English')}\n\n"
                f"📍 Entry: {entry}\n"
                f"🛑 SL: {sl}\n"
                f"🎯 TP1: {tp1}"
                + (f" | TP2: {tp2}" if tp2 != "—" else "")
                + (f" | TP3: {tp3}" if tp3 != "—" else "")
                + f"\n\nRR: {rr} | Basis: {basis}\n\n"
                f"Want early access? DM me — I'll send you the details.\n\n"
                f"#{sym} #ForexSignals #SMC #ICT #Gold #Trading"
            ),
        },

        # ── PERSIAN ──────────────────────────────────────────────────────
        "Persian": {
            "1": (
                f"🟢 سیگنال زنده | {sym} ({symbol_label(sym, 'Persian')})\n\n"
                f"📈 جهت: {direction_label(dirn, 'Persian')}\n"
                f"🎯 ورود: {entry}\n"
                f"🛑 حد ضرر: {sl}\n"
                f"{tp_lines_fa(tp1, tp2, tp3)}\n"
                f"⚡ ریسک: ۱٪ از حساب\n"
                f"📊 بر اساس: {basis}\n\n"
                f"💬 اگر می‌خوای هر روز سیگنال دریافت کنی — الان بهم پیام بده. فعلاً رایگانه.\n\n"
                f"#{sym} #سیگنال_فارکس #SMC #ICT #معامله‌گری"
            ),
            "2": (
                f"این ستاپ خیلی تمیزه — نمی‌تونم نشونش ندم.\n\n"
                f"🟢 {sym} — {'فرصت خرید' if dirn == 'BUY' else 'فرصت فروش'}\n\n"
                f"{'قیمت نقدینگی رو جارو کرد و روی اوردر بلاک واکنش داد.' if dirn == 'BUY' else 'قیمت نقدینگی بالا رو جارو کرد و روی اوردر بلاک نزولی واکنش داد.'} "
                f"بایاس روزانه {'صعودیه' if dirn == 'BUY' else 'نزولیه'}.\n\n"
                f"📍 ورود: {entry}\n"
                f"🛑 SL: {sl}\n"
                f"🎯 TP: {tp_compact_en(tp2, tp3)}\n\n"
                f"نسبت ریسک به ریوارد: {rr}\n\n"
                f"می‌خوای سیگنال‌هام رو زودتر دریافت کنی؟ بهم پیام بده.\n\n"
                f"#{sym} #سیگنال_فارکس #پرایس_اکشن #SMC #ICT"
            ),
            "3": (
                f"اعضای کانال تلگرامم این سیگنال رو قبل از این پست دریافت کردن.\n\n"
                f"🟢 {sym} — {direction_label(dirn, 'Persian')}\n\n"
                f"📍 ورود: {entry}\n"
                f"🛑 SL: {sl}\n"
                f"🎯 TP1: {tp1}"
                + (f" | TP2: {tp2}" if tp2 != "—" else "")
                + (f" | TP3: {tp3}" if tp3 != "—" else "")
                + f"\n\nریوارد: {rr} | بر اساس: {basis}\n\n"
                f"می‌خوای زودتر دسترسی داشته باشی؟ بهم پیام بده.\n\n"
                f"#{sym} #سیگنال_فارکس #SMC #ICT #طلا"
            ),
        },

        # ── RUSSIAN ──────────────────────────────────────────────────────
        "Russian": {
            "1": (
                f"🟢 ЖИВОЙ СИГНАЛ | {sym} ({symbol_label(sym, 'Russian')})\n\n"
                f"📈 Направление: {direction_label(dirn, 'Russian')}\n"
                f"🎯 Вход: {entry}\n"
                f"🛑 Стоп-лосс: {sl}\n"
                f"✅ TP1: {tp1}\n"
                + (f"✅ TP2: {tp2}\n" if tp2 != "—" else "")
                + (f"✅ TP3: {tp3}\n" if tp3 != "—" else "")
                + f"\n⚡ Риск: 1% от депозита\n"
                f"📊 Основа: {basis}\n\n"
                f"💬 Напиши в личку, если хочешь получать такие сигналы каждый день — пока бесплатно.\n\n"
                f"#{sym} #СигналыФорекс #SMC #ICT #Трейдинг"
            ),
            "2": (
                f"Этот сетап слишком чистый, чтобы молчать.\n\n"
                f"🟢 {sym} — {'возможность для покупки' if dirn == 'BUY' else 'возможность для продажи'}\n\n"
                f"{'Цена сняла ликвидность снизу и показала реакцию на ордер-блоке.' if dirn == 'BUY' else 'Цена сняла ликвидность сверху и показала реакцию на медвежьем ордер-блоке.'} "
                f"Дневной байас — {'бычий' if dirn == 'BUY' else 'медвежий'}.\n\n"
                f"📍 Вход: {entry}\n"
                f"🛑 SL: {sl}\n"
                f"🎯 TP: {tp_compact_en(tp2, tp3)}\n\n"
                f"Соотношение риск/прибыль: {rr}\n\n"
                f"Хочешь получать сигналы раньше? Напиши в личку.\n\n"
                f"#{sym} #СигналыФорекс #SMCTrading #ICT #PriceAction"
            ),
            "3": (
                f"Участники моего Telegram получили этот сигнал раньше этого поста.\n\n"
                f"🟢 {sym} — {direction_label(dirn, 'Russian')}\n\n"
                f"📍 Вход: {entry}\n"
                f"🛑 SL: {sl}\n"
                f"🎯 TP1: {tp1}"
                + (f" | TP2: {tp2}" if tp2 != "—" else "")
                + (f" | TP3: {tp3}" if tp3 != "—" else "")
                + f"\n\nRR: {rr} | Основа: {basis}\n\n"
                f"Хочешь ранний доступ? Напиши мне.\n\n"
                f"#{sym} #СигналыФорекс #SMC #ICT #Золото"
            ),
        },
    }
    return templates


# ─────────────────────────────────────────────
#  SELENIUM HELPERS
# ─────────────────────────────────────────────

def random_delay(min_s=None, max_s=None):
    a = min_s or MIN_DELAY_SECONDS
    b = max_s or MAX_DELAY_SECONDS
    wait = random.uniform(a, b)
    print(f"   ⏳ Waiting {int(wait)}s before next post...")
    time.sleep(wait)


def build_driver():
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--window-size=1280,900")
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(options=opts)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def load_session(driver):
    try:
        with open(SESSION_FILE, "r") as f:
            cookies = json.load(f)
        driver.get("https://www.facebook.com")
        time.sleep(2)
        for cookie in cookies:
            try:
                driver.add_cookie(cookie)
            except Exception:
                pass
        driver.refresh()
        time.sleep(3)
        return True
    except FileNotFoundError:
        return False


def save_session(driver):
    with open(SESSION_FILE, "w") as f:
        json.dump(driver.get_cookies(), f)


def manual_login(driver):
    if HEADLESS:
        raise RuntimeError(
            f"Facebook session is missing or expired: {SESSION_FILE}. "
            "Create the session interactively before enabling production posting."
        )
    driver.get("https://www.facebook.com")
    time.sleep(3)
    print()
    print("=" * 55)
    print("  Please log into Facebook in the browser window.")
    print("  Once logged in, come back here and press Enter.")
    print("=" * 55)
    input("  >>> Press Enter when you are logged in... ")
    save_session(driver)
    print("[✓] Session saved.")


def load_groups_from_excel():
    if not os.path.exists(GROUPS_FILE):
        return []
    wb = load_workbook(GROUPS_FILE)
    ws = wb.active
    groups = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        if len(row) > 5 and row[5] is False:
            continue
        num, name, url, language, template_num = row[0], row[1], row[2], row[3], row[4]
        if not url or not str(url).startswith("http"):
            continue
        lang = str(language).strip() if language else "English"
        tmpl = str(template_num).strip() if template_num else "1"
        groups.append({"num": num, "name": name, "url": url, "lang": lang, "tmpl": tmpl})
    return groups


def get_already_posted(signal_id):
    try:
        wb = load_workbook(LOG_FILE)
        ws = wb.active
        posted = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[3] == "✅ Success" and len(row) > 5 and row[5] == signal_id:
                posted.add(row[1])
        return posted
    except FileNotFoundError:
        return set()


def init_log():
    try:
        load_workbook(LOG_FILE)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Post Log"
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 35
        headers = ["#", "Group URL", "Posted At", "Status", "Notes", "Signal ID"]
        hf = PatternFill("solid", start_color="1F3864")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = hf
            cell.alignment = Alignment(horizontal="center")
        wb.save(LOG_FILE)


def log_result(row_num, url, status, signal_id, notes=""):
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    next_row = ws.max_row + 1
    fill_color = "E2EFDA" if status == "✅ Success" else "FCE4D6"
    fill = PatternFill("solid", start_color=fill_color)
    values = [row_num, url, datetime.now().strftime("%Y-%m-%d %H:%M"), status, notes, signal_id]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=c, value=v)
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill
    wb.save(LOG_FILE)


def post_to_group(driver, url, message):
    try:
        driver.get(url)
        time.sleep(random.uniform(4, 6))

        post_box = None
        for sel in [
            "//div[@role='button' and contains(., 'Write something')]",
            "//div[@aria-label='Write something...']",
            "//span[contains(text(), 'Write something')]",
        ]:
            try:
                post_box = WebDriverWait(driver, 6).until(
                    EC.element_to_be_clickable((By.XPATH, sel))
                )
                post_box.click()
                break
            except Exception:
                continue

        if not post_box:
            return False, "Could not find post box"

        time.sleep(random.uniform(1.5, 2.5))

        text_area = None
        for sel in [
            "//div[@contenteditable='true' and @role='textbox']",
            "//div[@aria-label='Write something...' and @contenteditable='true']",
            "//div[@data-lexical-editor='true']",
        ]:
            try:
                text_area = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, sel))
                )
                text_area.click()
                break
            except Exception:
                continue

        if not text_area:
            return False, "Could not find text area"

        time.sleep(1)

        driver.execute_script(
            """
            const el = arguments[0];
            el.focus();
            const dt = new DataTransfer();
            dt.setData('text/plain', arguments[1]);
            el.dispatchEvent(new ClipboardEvent('paste', {clipboardData: dt, bubbles: true}));
            """,
            text_area, message,
        )
        time.sleep(2)

        current_text = driver.execute_script("return arguments[0].innerText;", text_area)
        if not current_text.strip():
            text_area.send_keys(message)
            time.sleep(2)

        post_btn = None
        for sel in [
            "//div[@aria-label='Post' and @role='button']",
            "//button[contains(text(), 'Post')]",
            "//div[@role='button']//span[text()='Post']",
        ]:
            try:
                post_btn = WebDriverWait(driver, 8).until(
                    EC.element_to_be_clickable((By.XPATH, sel))
                )
                break
            except Exception:
                continue

        if not post_btn:
            return False, "Could not find Post button"

        time.sleep(random.uniform(0.5, 1.5))
        post_btn.click()
        time.sleep(4)
        return True, "Posted successfully"

    except Exception as e:
        return False, str(e)[:120]


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--signal-file", default=None)
    parser.add_argument("--preflight", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--preview", action="store_true")
    return parser.parse_args()


def preflight():
    checks = {
        "selenium": importlib.util.find_spec("selenium") is not None,
        "groups_file": os.path.isfile(GROUPS_FILE),
        "session_file": os.path.isfile(SESSION_FILE),
        "chrome": any(
            os.path.isfile(path)
            for path in ("/usr/bin/google-chrome", "/usr/bin/chromium", "/usr/bin/chromium-browser")
        ),
    }
    groups = load_groups_from_excel() if checks["groups_file"] else []
    checks["groups_count"] = len(groups)
    checks["ready"] = all(checks[k] for k in ("selenium", "groups_file", "session_file", "chrome")) and bool(groups)
    print(json.dumps(checks, ensure_ascii=False))
    return checks


def main():
    args = parse_args()
    if args.preview:
        sig = load_signal(args.signal_file)
        templates = build_templates(sig)
        print(json.dumps(templates, ensure_ascii=False))
        return 0
    checks = preflight()
    if args.preflight:
        return 0 if checks["ready"] else 2
    if not checks["ready"] and not args.dry_run:
        missing = [k for k in ("selenium", "groups_file", "session_file", "chrome") if not checks[k]]
        if not checks["groups_count"]:
            missing.append("configured_groups")
        raise RuntimeError(f"Facebook poster is not ready: {', '.join(missing)}")

    print("\n🚀 Facebook Group Auto Poster — Script 2 (Signal-Driven)")
    print("=" * 55)

    # Load live signal and build dynamic templates
    sig = load_signal(args.signal_file)
    signal_id = str(sig.get("signal_id") or "legacy")
    TEMPLATES = build_templates(sig)

    print(f"\n📊 Signal: {sig['symbol']} {sig['direction']} | Entry: {sig['entry']} | SL: {sig['sl']}")
    print(f"   TP1: {sig['tp1']} | TP2: {sig.get('tp2','—')} | TP3: {sig.get('tp3','—')} | RR: {sig.get('rr','—')}\n")

    groups = load_groups_from_excel()
    if not groups:
        raise RuntimeError(f"No Facebook groups found in {GROUPS_FILE}")

    already_posted = get_already_posted(signal_id)
    pending = [g for g in groups if g["url"] not in already_posted]

    print(f"[i] Total groups  : {len(groups)}")
    print(f"[i] Already posted: {len(already_posted)}")
    print(f"[i] Remaining     : {len(pending)}")
    print(f"[i] This batch    : {min(BATCH_SIZE, len(pending))}\n")

    if not pending:
        print("✅ All groups posted! Reset post_log.xlsx to start a new campaign.")
        return

    batch = pending[:BATCH_SIZE]
    init_log()

    if args.dry_run:
        for group in batch:
            lang = group["lang"]
            tmpl = group["tmpl"]
            message = TEMPLATES.get(lang, TEMPLATES["English"]).get(tmpl, TEMPLATES["English"]["1"])
            print(f"[DRY RUN] {group['url']} | {lang} #{tmpl} | {len(message)} chars")
        return 0

    driver = build_driver()
    try:
        session_loaded = load_session(driver)
        if not (session_loaded and "login" not in driver.current_url and "facebook.com" in driver.current_url):
            manual_login(driver)

        success_count = 0
        fail_count = 0

        for i, group in enumerate(batch, 1):
            lang = group["lang"]
            tmpl = group["tmpl"]
            message = TEMPLATES.get(lang, TEMPLATES["English"]).get(tmpl, TEMPLATES["English"]["1"])

            print(f"[{i}/{len(batch)}] {group['name'] or group['url']}")
            print(f"         Language: {lang} | Template: #{tmpl}")

            ok, notes = post_to_group(driver, group["url"], message)

            if ok:
                print(f"         ✅ Success")
                log_result(group["num"], group["url"], "✅ Success", signal_id, notes)
                success_count += 1
            else:
                print(f"         ❌ Failed: {notes}")
                log_result(group["num"], group["url"], "❌ Failed", signal_id, notes)
                fail_count += 1

            if i < len(batch):
                random_delay()

        print("\n" + "=" * 55)
        print(f"✅ Done: {success_count} posted, {fail_count} failed")
        remaining = len(pending) - len(batch)
        if remaining > 0:
            print(f"⏭  {remaining} groups remaining — run again tomorrow.")
        else:
            print("🎉 All groups posted!")

    finally:
        driver.quit()


if __name__ == "__main__":
    if "--preflight" in sys.argv:
        raise SystemExit(main())
    os.makedirs(os.path.dirname(os.path.abspath(LOCK_FILE)), exist_ok=True)
    with open(LOCK_FILE, "w", encoding="utf-8") as lock:
        fcntl.flock(lock, fcntl.LOCK_EX)
        raise SystemExit(main())
