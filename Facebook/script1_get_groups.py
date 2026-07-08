"""
SCRIPT 1 — Facebook Group URL Scraper
======================================
Logs into Facebook using your credentials, navigates to your joined groups,
scrapes all group names and URLs, and saves them to an Excel file.

REQUIREMENTS:
    pip install selenium openpyxl webdriver-manager

USAGE:
    1. Fill in your FB_EMAIL and FB_PASSWORD below
    2. Run: python script1_get_groups.py
    3. Output: fb_my_groups.xlsx (in the same folder)
"""

import time
import random
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
#  YOUR FACEBOOK CREDENTIALS
# ─────────────────────────────────────────────
FB_EMAIL    = "your_email@example.com"   # ← Replace
FB_PASSWORD = "your_password"            # ← Replace

# ─────────────────────────────────────────────
#  SETTINGS
# ─────────────────────────────────────────────
OUTPUT_FILE        = "fb_my_groups.xlsx"
SCROLL_PAUSE       = 2.5   # seconds between scrolls
MAX_SCROLL_ROUNDS  = 30    # how many times to scroll (increase if you have many groups)
SESSION_FILE       = "fb_session.json"   # saves cookies so you don't login every time


def random_delay(min_s=1.5, max_s=3.5):
    time.sleep(random.uniform(min_s, max_s))


def build_driver(headless=False):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--window-size=1280,900")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts
    )
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def load_session(driver):
    """Load saved cookies to skip login."""
    try:
        with open(SESSION_FILE, "r") as f:
            cookies = json.load(f)
        driver.get("https://www.facebook.com")
        time.sleep(2)
        for cookie in cookies:
            try:
                driver.add_cookie(cookie)
            except Exception:
                pass
        driver.refresh()
        time.sleep(3)
        return True
    except FileNotFoundError:
        return False


def save_session(driver):
    cookies = driver.get_cookies()
    with open(SESSION_FILE, "w") as f:
        json.dump(cookies, f)
    print("[✓] Session saved — next run will skip login.")


def dismiss_popups(driver):
    """Try to close any cookie/consent popups Facebook may show."""
    popup_xpaths = [
        "//button[contains(text(), 'Allow all cookies')]",
        "//button[contains(text(), 'Accept all')]",
        "//button[contains(text(), 'Allow essential')]",
        "//button[contains(text(), 'Only allow essential cookies')]",
        "//button[contains(text(), 'Decline optional cookies')]",
        "//div[@aria-label='Allow all cookies']",
        "//div[@aria-label='Accept all']",
        "//div[@role='button' and contains(., 'Accept')]",
        "//div[@role='button' and contains(., 'Allow')]",
    ]
    for xpath in popup_xpaths:
        try:
            btn = driver.find_element(By.XPATH, xpath)
            btn.click()
            print("   [i] Dismissed a cookie/consent popup.")
            time.sleep(2)
            return
        except Exception:
            continue


def login(driver):
    print("[→] Logging in to Facebook...")
    driver.get("https://www.facebook.com")
    time.sleep(5)  # wait longer for page + any popups

    # Dismiss cookie popups BEFORE trying to find the login form
    dismiss_popups(driver)
    time.sleep(2)

    # If we're not already on login page, go there
    if "login" not in driver.current_url:
        driver.get("https://www.facebook.com/login")
        time.sleep(4)
        dismiss_popups(driver)
        time.sleep(2)

    print("   [i] Current URL:", driver.current_url)

    # Wait up to 30 seconds for email field
    wait = WebDriverWait(driver, 30)
    try:
        email_field = wait.until(EC.presence_of_element_located((By.ID, "email")))
    except Exception:
        # Try alternative selectors
        try:
            email_field = driver.find_element(By.NAME, "email")
        except Exception:
            try:
                email_field = driver.find_element(By.XPATH, "//input[@type='email']")
            except Exception:
                raise RuntimeError(
                    "Could not find the email field. Facebook may have shown a popup.\n"
                    "The browser window is open — please dismiss any popups manually,\n"
                    "then complete the login yourself. The script will continue."
                )

    email_field.clear()
    time.sleep(0.5)
    for ch in FB_EMAIL:
        email_field.send_keys(ch)
        time.sleep(random.uniform(0.05, 0.15))

    time.sleep(0.5)

    try:
        pass_field = driver.find_element(By.ID, "pass")
    except Exception:
        pass_field = driver.find_element(By.NAME, "pass")

    pass_field.clear()
    time.sleep(0.3)
    for ch in FB_PASSWORD:
        pass_field.send_keys(ch)
        time.sleep(random.uniform(0.05, 0.15))

    time.sleep(random.uniform(0.8, 1.5))

    try:
        driver.find_element(By.ID, "loginbutton").click()
    except Exception:
        try:
            driver.find_element(By.XPATH, "//button[@type='submit']").click()
        except Exception:
            pass_field.send_keys(Keys.RETURN)

    print("   [i] Login submitted — waiting for page to load...")
    time.sleep(8)

    # If 2FA or checkpoint appears, give user time to complete it manually
    for attempt in range(12):  # wait up to 60 more seconds
        url = driver.current_url
        if "login" not in url and "checkpoint" not in url and "two_step" not in url:
            break
        if attempt == 3:
            print("   [!] If you see a 2FA or verification prompt in the browser,")
            print("       please complete it manually. Waiting up to 60 seconds...")
        time.sleep(5)
    else:
        raise RuntimeError("Login failed or 2FA not completed in time.")

    print("[✓] Logged in successfully.")
    save_session(driver)


def scrape_groups(driver):
    """Navigate to joined groups page and collect all group links."""
    print("[→] Navigating to your groups...")
    driver.get("https://www.facebook.com/groups/joins/")
    time.sleep(4)

    groups = {}   # url -> name  (dict deduplicates)

    print("[→] Scrolling to load all groups...")
    last_count = 0
    no_change_rounds = 0

    for scroll_round in range(MAX_SCROLL_ROUNDS):
        # Collect all group links visible so far
        anchors = driver.find_elements(By.XPATH, "//a[contains(@href, '/groups/')]")
        for a in anchors:
            href = a.get_attribute("href") or ""
            # Filter to actual group pages (not feed, discover, etc.)
            if "/groups/" in href and not any(
                x in href for x in [
                    "/groups/feed", "/groups/discover", "/groups/joins",
                    "/groups/create", "?", "#", "members", "media", "events", "files"
                ]
            ):
                # Clean URL: strip trailing slashes and query params
                clean = href.split("?")[0].rstrip("/")
                name = a.text.strip() or clean.split("/groups/")[-1].replace("-", " ").title()
                if clean not in groups and len(clean) > 30:
                    groups[clean] = name

        current_count = len(groups)
        print(f"   Round {scroll_round + 1}: {current_count} groups found so far...")

        if current_count == last_count:
            no_change_rounds += 1
            if no_change_rounds >= 4:
                print("[✓] No new groups found after 4 rounds — scraping complete.")
                break
        else:
            no_change_rounds = 0

        last_count = current_count

        # Scroll down
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE)

    print(f"[✓] Total groups found: {len(groups)}")
    return groups


def detect_language(text):
    """Detect language from group name using Unicode character ranges."""
    if not text:
        return "English"
    persian_arabic = 0
    cyrillic = 0
    for ch in text:
        cp = ord(ch)
        # Persian / Arabic Unicode block: U+0600–U+06FF and U+FB50–U+FDFF
        if (0x0600 <= cp <= 0x06FF) or (0xFB50 <= cp <= 0xFDFF):
            persian_arabic += 1
        # Cyrillic Unicode block: U+0400–U+04FF
        elif 0x0400 <= cp <= 0x04FF:
            cyrillic += 1
    if persian_arabic > 0:
        return "Persian"
    elif cyrillic > 0:
        return "Russian"
    else:
        return "English"


def save_to_excel(groups: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "My Facebook Groups"

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30

    # Header
    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    headers = ["#", "Group Name", "Group URL", "Language", "Message Template #"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Color fills per language
    body_font    = Font(name="Arial", size=10)
    fill_english = PatternFill("solid", start_color="E2EFDA")   # green
    fill_persian = PatternFill("solid", start_color="FCE4D6")   # orange
    fill_russian = PatternFill("solid", start_color="EAD1DC")   # pink
    fill_default = PatternFill("solid", start_color="FFFFFF")

    lang_fills = {
        "English": fill_english,
        "Persian": fill_persian,
        "Russian": fill_russian,
    }

    # Template rotation counters per language
    template_counters = {"English": 0, "Persian": 0, "Russian": 0}

    for i, (url, name) in enumerate(groups.items(), 1):
        language = detect_language(name)
        # Rotate templates 1 → 2 → 3 → 1 → 2 → 3 per language
        template_counters[language] += 1
        template_num = str((template_counters[language] - 1) % 3 + 1)

        fill = lang_fills.get(language, fill_default)
        row_data = [i, name, url, language, template_num]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))
        ws.row_dimensions[i + 1].height = 18

    # Legend
    last_row = len(groups) + 3
    legend = ws.cell(row=last_row, column=1,
        value="🟩 Green = English   🟧 Orange = Persian   🟪 Pink = Russian  |  "
              "Language & Template # are auto-detected. You can manually correct any row if needed.")
    legend.font = Font(italic=True, color="444444", name="Arial", size=9)

    wb.save(OUTPUT_FILE)
    print(f"[✓] Saved to {OUTPUT_FILE}")

    # Print summary
    counts = {"English": 0, "Persian": 0, "Russian": 0}
    for name in groups.values():
        counts[detect_language(name)] += 1
    print(f"   Language breakdown → English: {counts['English']} | Persian: {counts['Persian']} | Russian: {counts['Russian']}")


def manual_login(driver):
    """Open Facebook and wait for the user to log in manually."""
    driver.get("https://www.facebook.com")
    time.sleep(3)

    print()
    print("=" * 55)
    print("  A Chrome window has opened.")
    print("  Please log into Facebook manually in that window.")
    print("  Once you are fully logged in and can see your")
    print("  Facebook homepage, come back here and press Enter.")
    print("=" * 55)
    input("  >>> Press Enter when you are logged in... ")
    print()

    # Save session so future runs skip this step
    save_session(driver)
    print("[✓] Session saved — future runs will skip manual login.")


def main():
    driver = build_driver(headless=False)
    try:
        # Try to reuse saved session first
        session_loaded = load_session(driver)
        if session_loaded and "login" not in driver.current_url and "facebook.com" in driver.current_url:
            print("[✓] Reusing saved session — skipping login.")
        else:
            manual_login(driver)

        groups = scrape_groups(driver)

        if not groups:
            print("[!] No groups found. Facebook may have changed its layout.")
            print("    Check the Chrome window to see what page is showing.")
            return

        save_to_excel(groups)
        print(f"\n✅ Done! Open '{OUTPUT_FILE}' and fill in the Language and Template # columns.")
        print("   Then run script2_post_to_groups.py to start posting.\n")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
